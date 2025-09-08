from typing import Any, TypeAlias, IO, TYPE_CHECKING

import openpyxl as op

from ut_log.log import LogEq
from ut_obj.io import Io

from ut_xls.op.wb import Wb
from ut_xls.op.ws import Ws

if TYPE_CHECKING:
    from os import PathLike
    from _typeshed import SupportsRead
    TyOpFileSrc = str | PathLike[str] | IO[bytes] | SupportsRead[bytes]
else:
    from os import PathLike
    TyOpFileSrc = str | PathLike[str] | IO[bytes]

TyWb: TypeAlias = op.workbook.workbook.Workbook
TyWs: TypeAlias = op.worksheet.worksheet.Worksheet
TyOpWb: TypeAlias = op.Workbook

TyArr = list[Any]
TyDic = dict[Any, Any]
TyAoA = list[TyArr]
TyAoD = list[TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoWs = dict[Any, TyWs]
TyAoD_DoAoD = TyAoD | TyDoAoD
TyPath = str
TyPathnm = str
TyStr = str

TySheet = int | str
TySheets = int | str | list[int | str]
TySheetname = str
TySheetnames = list[TySheetname]

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnAoD_DoAoD = None | TyAoD_DoAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnDoAoD = None | TyDoAoD
TnDoWs = None | TyDoWs
TnSheet = None | TySheet
TnSheets = None | TySheets
TnSheetname = None | TySheetname
TnSheetnames = None | TySheetnames
TnWs = None | TyWs
# TnDf_DoDf = TnPdDf_DoPdDf | TnPlDf_DoPlDf
TnOpWb = None | TyOpWb


class IoiPathWb:

    @staticmethod
    def load(io: TyOpFileSrc, **kwargs) -> TyWb:
        if io == '':
            raise Exception('io is empty String')
        if io is None:
            raise Exception('io is None')
        try:
            wb: TyWb = op.load_workbook(io, **kwargs)
        except Exception as e:
            msg = f"openpyxl.load_workbook for io = {io!r} throw exception {e}"
            raise Exception(msg)
        return wb

    @classmethod
    def read_wb_to_aod(
            cls, io: TyOpFileSrc, sheet: TnSheet, **kwargs) -> TyAoD:
        Io.verify(io)
        _wb: TyWb = cls.load(io, **kwargs)
        _aod: TyAoD = Wb.to_aod(_wb, sheet)
        return _aod

    @classmethod
    def read_wb_to_doaod(
            cls, io: TyOpFileSrc, sheet: TnSheets, **kwargs) -> TyDoAoD:
        Io.verify(io)
        _wb: TyWb = cls.load(io, **kwargs)
        _doaod: TyDoAoD = Wb.to_doaod(_wb, sheet)
        return _doaod

    @classmethod
    def read_wb_to_aod_or_doaod(
            cls, io: TyOpFileSrc, sheet: TnSheets, **kwargs) -> TnAoD_DoAoD:
        Io.verify(io)
        _wb: TyWb = cls.load(io, **kwargs)
        _aod_doaod: TnAoD_DoAoD = Wb.to_aod_or_doaod(_wb, sheet)
        return _aod_doaod

    @classmethod
    def read_wb_to_aoa(
            cls, io: TyOpFileSrc, **kwargs) -> tuple[TyAoA, TyAoA]:
        Io.verify(io)
        wb: TyWb = cls.load(io)
        heads_sheet_name = kwargs.get('headers_sheet_name')
        ws_names: TySheetnames = Wb.sh_sheetnames(wb, **kwargs)
        aoa: TyAoA = []
        if heads_sheet_name is not None:
            ws = wb[heads_sheet_name]
            heads: TyAoA = Ws.sh_headers(ws, **kwargs)
        else:
            heads = []
        for ws_name in ws_names:
            LogEq.debug("ws_name", ws_name)
            ws = wb[ws_name]
            aoa_ws = Ws.sh_aoa(ws, sheet_name=ws_name, **kwargs)
            aoa.extend(aoa_ws)
            LogEq.debug("aoa_ws", aoa_ws)
        return heads, aoa

    @classmethod
    def read_wb_to_aoa_by_prefix(cls, **kwargs) -> TyAoA:
        # ex_read_workbook_2_aoa(cls, **kwargs):
        # def ex_read_aoa(cls, **kwargs):
        prefix = kwargs.get('prefix')
        if prefix is not None:
            prefix = f"_{prefix}"
        in_io: TyOpFileSrc = kwargs.get(f'in_path{prefix}', '')
        row_start = kwargs.get(f'row_start{prefix}')
        cols_count = kwargs.get(f'cols_count{prefix}')
        sw_add_sheet_name = kwargs.get(f'sw_add_sheet_name{prefix}')
        sheet_names = kwargs.get(f'sheet_names{prefix}')
        headers_sheet_name = kwargs.get(f'headers_sheet_name{prefix}')
        headers_start = kwargs.get(f'headers_start{prefix}')
        Io.verify(in_io)
        heads, aoa = cls.read_wb_to_aoa(
                in_io,
                row_start=row_start,
                cols_count=cols_count,
                sw_add_sheet_name=sw_add_sheet_name,
                sheet_names=sheet_names,
                headers_sheet_name=headers_sheet_name,
                headers_start=headers_start)
        return aoa

    @classmethod
    def sh_wb_adm(
            cls, in_path: TyPath, aod_adm: TnAoD, in_sheet_adm: TyStr
    ) -> TnOpWb:
        """
        Administration processsing for evup xlsx workbooks
        """
        _wb: TnOpWb = cls.load(in_path, read_only=False)
        Wb.update_wb_with_aod(_wb, aod_adm, in_sheet_adm)
        return _wb

    @classmethod
    def sh_wb_del(
            cls, in_path: TyPath, aod_del: TnAoD, in_sheet_del: TyStr
    ) -> TnOpWb:
        """
        Delete processsing for evup xlsx workbooks
        """
        _wb: TnOpWb = cls.load(in_path, read_only=False)
        Wb.update_wb_with_aod(_wb, aod_del, in_sheet_del)
        return _wb

    @classmethod
    def sh_wb_reg(
            cls,
            in_path: TyPath,
            aod_adm: TnAoD, aod_del: TnAoD,
            in_sheet_adm: TyStr, in_sheet_del: TyStr
    ) -> TnOpWb:
        """
        Delete processsing for evup xlsx workbooks
        """
        _wb: TnOpWb = cls.load(in_path, read_only=False)
        Wb.update_wb_with_aod(_wb, aod_adm, in_sheet_adm)
        Wb.update_wb_with_aod(_wb, aod_del, in_sheet_del)
        return _wb
