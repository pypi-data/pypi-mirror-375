import asyncio
import json
import os
import shutil
from datetime import datetime
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any, Awaitable, Dict, List, Optional, TypeVar

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import AfterValidator, BaseModel, field_validator
from pydantic_core import PydanticCustomError

from synapse_sdk.clients.exceptions import ClientError
from synapse_sdk.clients.utils import get_batched_list
from synapse_sdk.clients.validators.collections import FileSpecificationValidator
from synapse_sdk.i18n import gettext as _
from synapse_sdk.plugins.categories.base import Action
from synapse_sdk.plugins.categories.decorators import register_action
from synapse_sdk.plugins.enums import PluginCategory, RunMethod
from synapse_sdk.plugins.exceptions import ActionError
from synapse_sdk.plugins.models import Run
from synapse_sdk.shared.enums import Context
from synapse_sdk.utils.pydantic.validators import non_blank
from synapse_sdk.utils.storage import get_pathlib

# Type variable for generic async return type
T = TypeVar('T')


class PathAwareJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder that handles Path-like objects."""

    def default(self, obj):
        if hasattr(obj, '__fspath__') or hasattr(obj, 'as_posix'):
            # Handle Path-like objects (including UPath, SFTPPath, pathlib.Path, etc.)
            return str(obj)
        elif hasattr(obj, 'isoformat'):
            # Handle datetime objects
            return obj.isoformat()
        # Let the base class handle other types
        return super().default(obj)


class UploadStatus(str, Enum):
    SUCCESS = 'success'
    FAILED = 'failed'


class UploadRun(Run):
    class UploadEventLog(BaseModel):
        """Upload event log model."""

        info: Optional[str] = None
        status: Context
        created: str

    class DataFileLog(BaseModel):
        """Data file log model."""

        data_file_info: str | None
        status: UploadStatus
        created: str

    class DataUnitLog(BaseModel):
        """Data unit log model."""

        data_unit_id: int | None
        status: UploadStatus
        created: str
        data_unit_meta: dict | None

    class TaskLog(BaseModel):
        """Task log model."""

        task_id: int | None
        status: UploadStatus
        created: str

    class MetricsRecord(BaseModel):
        """Metrics record model."""

        stand_by: int
        failed: int
        success: int

    LOG_MESSAGES = {
        # Validation errors - show in both log_message and EventLog
        'STORAGE_VALIDATION_FAILED': {
            'message': 'Storage validation failed.',
            'level': Context.DANGER,
        },
        'COLLECTION_VALIDATION_FAILED': {
            'message': 'Collection validation failed.',
            'level': Context.DANGER,
        },
        'PROJECT_VALIDATION_FAILED': {
            'message': 'Project validation failed.',
            'level': Context.DANGER,
        },
        'VALIDATION_FAILED': {
            'message': 'Validation failed.',
            'level': Context.DANGER,
        },
        'NO_FILES_FOUND': {
            'message': 'Files not found on the path.',
            'level': Context.WARNING,
        },
        'NO_FILES_UPLOADED': {
            'message': 'No files were uploaded.',
            'level': Context.WARNING,
        },
        'NO_DATA_UNITS_GENERATED': {
            'message': 'No data units were generated.',
            'level': Context.WARNING,
        },
        'NO_TYPE_DIRECTORIES': {
            'message': 'No type-based directory structure found.',
            'level': Context.INFO,
        },
        'EXCEL_SECURITY_VIOLATION': {
            'message': 'Excel security validation failed: {}',
            'level': Context.DANGER,
        },
        'EXCEL_PARSING_ERROR': {
            'message': 'Excel parsing failed: {}',
            'level': Context.DANGER,
        },
        'EXCEL_METADATA_LOADED': {
            'message': 'Excel metadata loaded for {} files',
            'level': None,
        },
        'UPLOADING_DATA_FILES': {
            'message': 'Uploading data files...',
            'level': None,
        },
        'GENERATING_DATA_UNITS': {
            'message': 'Generating data units...',
            'level': None,
        },
        'IMPORT_COMPLETED': {
            'message': 'Import completed.',
            'level': None,
        },
        'TYPE_DIRECTORIES_FOUND': {
            'message': 'Found type directories: {}',
            'level': None,
        },
        'TYPE_STRUCTURE_DETECTED': {
            'message': 'Detected type-based directory structure',
            'level': None,
        },
        'FILES_DISCOVERED': {
            'message': 'Discovered {} files',
            'level': None,
        },
        'NO_FILES_FOUND_WARNING': {
            'message': 'No files found.',
            'level': Context.WARNING,
        },
        'FILE_UPLOAD_FAILED': {
            'message': 'Failed to upload file: {}',
            'level': Context.DANGER,
        },
        'DATA_UNIT_BATCH_FAILED': {
            'message': 'Failed to create data units batch: {}',
            'level': Context.DANGER,
        },
        'FILENAME_TOO_LONG': {
            'message': 'Skipping file with overly long name: {}...',
            'level': Context.WARNING,
        },
        'MISSING_REQUIRED_FILES': {
            'message': '{} missing required files: {}',
            'level': Context.WARNING,
        },
        'EXCEL_FILE_NOT_FOUND': {
            'message': 'Excel metadata file not found: {}',
            'level': Context.WARNING,
        },
        # Debug information - only for EventLog
        'EXCEL_FILE_VALIDATION_STARTED': {
            'message': 'Excel file validation started',
            'level': Context.INFO,
        },
        'EXCEL_WORKBOOK_LOADED': {
            'message': 'Excel workbook loaded successfully',
            'level': Context.INFO,
        },
        'FILE_ORGANIZATION_STARTED': {
            'message': 'File organization started',
            'level': Context.INFO,
        },
        'BATCH_PROCESSING_STARTED': {
            'message': 'Batch processing started: {} batches of {} items each',
            'level': Context.INFO,
        },
        'EXCEL_SECURITY_VALIDATION_STARTED': {
            'message': 'Excel security validation started for file size: {} bytes',
            'level': Context.INFO,
        },
        'EXCEL_MEMORY_ESTIMATION': {
            'message': 'Excel memory estimation: {} bytes (file) * 3 = {} bytes (estimated)',
            'level': Context.INFO,
        },
        'EXCEL_FILE_NOT_FOUND_PATH': {
            'message': 'Excel metadata file not found',
            'level': Context.WARNING,
        },
        'EXCEL_SECURITY_VALIDATION_FAILED': {
            'message': 'Excel security validation failed: {}',
            'level': Context.DANGER,
        },
        'EXCEL_PARSING_FAILED': {
            'message': 'Excel parsing failed: {}',
            'level': Context.DANGER,
        },
        'EXCEL_INVALID_FILE_FORMAT': {
            'message': 'Invalid Excel file format: {}',
            'level': Context.DANGER,
        },
        'EXCEL_FILE_TOO_LARGE': {
            'message': 'Excel file too large to process (memory limit exceeded)',
            'level': Context.DANGER,
        },
        'EXCEL_FILE_ACCESS_ERROR': {
            'message': 'File access error reading excel metadata: {}',
            'level': Context.DANGER,
        },
        'EXCEL_UNEXPECTED_ERROR': {
            'message': 'Unexpected error reading excel metadata: {}',
            'level': Context.DANGER,
        },
    }

    def log_message_with_code(self, code: str, *args, level: Optional[Context] = None):
        """Unified logging method that handles both log_message and EventLog based on configuration."""
        if code not in self.LOG_MESSAGES:
            self.log_message(f'Unknown log code: {code}')
            self.log_upload_event('UNKNOWN_LOG_CODE', code)
            return

        log_config = self.LOG_MESSAGES[code]
        message = log_config['message'].format(*args) if args else log_config['message']
        log_level = level or log_config['level'] or Context.INFO

        # Log to message if configured
        if log_level == Context.INFO.value:
            self.log_message(message, context=log_level.value)
        else:
            self.log_upload_event(code, *args, level)

    def log_upload_event(self, code: str, *args, level: Optional[Context] = None):
        """Log upload event using predefined code."""
        if code not in self.LOG_MESSAGES:
            now = datetime.now().isoformat()
            self.log(
                'upload_event',
                self.UploadEventLog(info=f'Unknown log code: {code}', status=Context.DANGER, created=now).model_dump(),
            )
            return

        log_config = self.LOG_MESSAGES[code]
        message = log_config['message'].format(*args) if args else log_config['message']
        log_level = level or log_config['level'] or Context.INFO

        now = datetime.now().isoformat()
        self.log(
            'upload_event',
            self.UploadEventLog(info=message, status=log_level, created=now).model_dump(),
        )

    def log_data_file(self, data_file_info: dict, status: UploadStatus):
        """Upload data_file log.

        Args:
            data_file_info (dict): The json info of the data file.
            status (UploadStatus): The status of the data file.
        """
        now = datetime.now().isoformat()
        # Use custom JSON encoder to handle Path-like objects
        data_file_info_str = json.dumps(data_file_info, ensure_ascii=False, cls=PathAwareJSONEncoder)
        self.log(
            'upload_data_file',
            self.DataFileLog(data_file_info=data_file_info_str, status=status.value, created=now).model_dump(),
        )

    def log_data_unit(self, data_unit_id: int, status: UploadStatus, data_unit_meta: dict | None = None):
        """Upload data_unit log.

        Args:
            data_unit_id (int): The ID of the data unit.
            status (UploadStatus): The status of the data unit.
            data_unit_meta (dict | None): The metadata of the data unit.
        """
        now = datetime.now().isoformat()
        self.log(
            'upload_data_unit',
            self.DataUnitLog(
                data_unit_id=data_unit_id, status=status.value, created=now, data_unit_meta=data_unit_meta
            ).model_dump(),
        )

    def log_task(self, task_id: int, status: UploadStatus):
        """Upload task log.

        Args:
            task_id (int): The ID of the task.
            status (UploadStatus): The status of the task.
        """
        now = datetime.now().isoformat()
        self.log('upload_task', self.TaskLog(task_id=task_id, status=status.value, created=now).model_dump())

    def log_metrics(self, record: MetricsRecord, category: str):
        """Log upload metrics.
        Args:
            record (MetricsRecord): The metrics record to log.
            category (str): The category of the metrics.
        """
        record = self.MetricsRecord.model_validate(record)
        self.set_metrics(value=record.model_dump(), category=category)


class ExcelSecurityError(Exception):
    """Custom exception for Excel security validation errors."""

    pass


class ExcelParsingError(Exception):
    """Custom exception for Excel parsing errors."""

    pass


class ExcelSecurityConfig:
    """Configuration class for Excel security settings."""

    def __init__(self):
        # File size limits
        self.MAX_FILE_SIZE_MB = int(os.getenv('EXCEL_MAX_FILE_SIZE_MB', '10'))
        self.MAX_FILE_SIZE_BYTES = self.MAX_FILE_SIZE_MB * 1024 * 1024

        # Memory limits
        self.MAX_MEMORY_USAGE_MB = int(os.getenv('EXCEL_MAX_MEMORY_MB', '30'))
        self.MAX_MEMORY_USAGE_BYTES = self.MAX_MEMORY_USAGE_MB * 1024 * 1024

        # Content limits
        self.MAX_ROWS = int(os.getenv('EXCEL_MAX_ROWS', '10000'))
        self.MAX_COLUMNS = int(os.getenv('EXCEL_MAX_COLUMNS', '50'))

        # String length limits
        self.MAX_FILENAME_LENGTH = int(os.getenv('EXCEL_MAX_FILENAME_LENGTH', '255'))
        self.MAX_COLUMN_NAME_LENGTH = int(os.getenv('EXCEL_MAX_COLUMN_NAME_LENGTH', '100'))
        self.MAX_METADATA_VALUE_LENGTH = int(os.getenv('EXCEL_MAX_METADATA_VALUE_LENGTH', '1000'))


class ExcelMetadataUtils:
    """Utility class for Excel metadata processing with shared validation logic."""

    def __init__(self, config: ExcelSecurityConfig):
        self.config = config

    def validate_and_truncate_string(self, value: str, max_length: int) -> str:
        """Validate and truncate string to specified maximum length.

        Args:
            value: String value to validate and truncate
            max_length: Maximum allowed length

        Returns:
            str: Validated and potentially truncated string
        """
        if not isinstance(value, str):
            value = str(value)

        value = value.strip()

        if len(value) > max_length:
            return value[:max_length]

        return value

    def is_valid_filename_length(self, filename: str) -> bool:
        """Check if filename length is within acceptable limits.

        Args:
            filename: Filename to check

        Returns:
            bool: True if filename length is acceptable
        """
        return len(filename.strip()) <= self.config.MAX_FILENAME_LENGTH


class UploadParams(BaseModel):
    """Upload action parameters.

    This class defines all configuration parameters for the upload action, including
    advanced asynchronous upload capabilities for improved performance.

    Args:
        name (str): The name of the action.
        description (str | None): The description of the action.
        checkpoint (int | None): The checkpoint of the action.
        path (str): The path of the action.
        storage (int): The storage of the action.
        collection (int): The collection of the action.
        project (int | None): The project of the action.
        excel_metadata_path (str | None): Path to excel file containing metadata.
            Defaults to 'meta.xlsx' or 'meta.xls' in the path directory.
        is_recursive (bool): Enable recursive file discovery in subdirectories. Defaults to False.
        use_async_upload (bool): Enable asynchronous upload data file processing for improved performance.
            Defaults to True.
        max_file_size_mb (int): The maximum file size not using chunked_upload in MB. Defaults to 50MB.
        creating_data_unit_batch_size (int): The batch size for creating data units. Defaults to 100.
        extra_params (dict | None): Extra parameters for the action.
            Example: {"include_metadata": True, "compression": "gzip"}

    Note:
        Async upload requires plugin developers to implement handle_upload_files_async()
        method for maximum benefit. Default implementation provides compatibility
        but limited performance improvement.
    """

    name: Annotated[str, AfterValidator(non_blank)]
    description: str | None
    path: str
    storage: int
    collection: int
    project: int | None
    excel_metadata_path: str | None = None
    is_recursive: bool = True
    max_file_size_mb: int = 50
    creating_data_unit_batch_size: int = 1
    use_async_upload: bool = True
    extra_params: dict | None = None

    @field_validator('storage', mode='before')
    @classmethod
    def check_storage_exists(cls, value: str, info) -> str:
        """Validate synapse-backend storage exists.

        TODO: Need to define validation method naming convention.
        TODO: Need to make validation method reusable.
        """
        action = info.context['action']
        client = action.client
        try:
            client.get_storage(value)
        except ClientError:
            raise PydanticCustomError('client_error', _('Error occurred while checking storage exists.'))
        return value

    @field_validator('collection', mode='before')
    @classmethod
    def check_collection_exists(cls, value: str, info) -> str:
        """Validate synapse-backend collection exists."""
        action = info.context['action']
        client = action.client
        try:
            client.get_data_collection(value)
        except ClientError:
            raise PydanticCustomError('client_error', _('Error occurred while checking collection exists.'))
        return value

    @field_validator('project', mode='before')
    @classmethod
    def check_project_exists(cls, value: str, info) -> str:
        """Validate synapse-backend project exists."""
        if not value:
            return value

        action = info.context['action']
        client = action.client
        try:
            client.get_project(value)
        except ClientError:
            raise PydanticCustomError('client_error', _('Error occurred while checking project exists.'))
        return value

    @field_validator('excel_metadata_path', mode='before')
    @classmethod
    def check_excel_metadata_path(cls, value: str, info) -> str:
        """Validate excel metadata file exists and is secure if provided.

        This validator performs comprehensive security checks including:
        - File existence and format validation
        - File size limits (max 10MB)
        - Basic security checks for file content

        Args:
            value: The excel file path to validate
            info: Validation context information

        Returns:
            str: The validated file path

        Raises:
            PydanticCustomError: If validation fails
        """
        if not value:
            return value

        excel_path = Path(value)

        # Check file existence
        if not excel_path.exists():
            raise PydanticCustomError('file_not_found', _('Excel metadata file not found.'))

        # Check file extension
        if excel_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise PydanticCustomError('invalid_file_type', _('Excel metadata file must be .xlsx or .xls format.'))

        # Security check: file size limit
        file_size = excel_path.stat().st_size
        excel_config = ExcelSecurityConfig()
        if file_size > excel_config.MAX_FILE_SIZE_BYTES:
            raise PydanticCustomError(
                'file_too_large',
                _('Excel metadata file is too large. Maximum size is {}MB.').format(excel_config.MAX_FILE_SIZE_MB),
            )

        # Basic security check: ensure file is readable and not corrupted
        try:
            with open(excel_path, 'rb') as f:
                # Read first few bytes to check if it's a valid Excel file
                header = f.read(8)
                if not header:
                    raise PydanticCustomError('invalid_file', _('Excel metadata file appears to be empty.'))

                # Check for valid Excel file signatures
                if excel_path.suffix.lower() == '.xlsx':
                    # XLSX files start with PK (ZIP signature)
                    if not header.startswith(b'PK'):
                        raise PydanticCustomError('invalid_file', _('Excel metadata file appears to be corrupted.'))
                elif excel_path.suffix.lower() == '.xls':
                    # XLS files have specific OLE signatures
                    if not (header.startswith(b'\xd0\xcf\x11\xe0') or header.startswith(b'\x09\x08')):
                        raise PydanticCustomError('invalid_file', _('Excel metadata file appears to be corrupted.'))

        except (OSError, IOError):
            raise PydanticCustomError('file_access_error', _('Cannot access Excel metadata file.'))

        return value


@register_action
class UploadAction(Action):
    """Upload action class.

    Attrs:
        name (str): The name of the action.
        category (PluginCategory): The category of the action.
        method (RunMethod): The method to run of the action.

    Progress Categories:
        analyze_collection: The progress category for the analyze collection process.
        data_file_upload: The progress category for the upload process.
        generate_data_units: The progress category for the generate data units process.

    Metrics Categories:
        data_file: The metrics category for the data file.
        data_unit: The metrics category for the data unit.
    """

    name = 'upload'
    category = PluginCategory.UPLOAD
    method = RunMethod.JOB
    run_class = UploadRun
    progress_categories = {
        'analyze_collection': {
            'proportion': 2,
        },
        'upload_data_files': {
            'proportion': 38,
        },
        'generate_data_units': {
            'proportion': 60,
        },
    }
    metrics_categories = {
        'data_files': {
            'stand_by': 0,
            'failed': 0,
            'success': 0,
        },
        'data_units': {
            'stand_by': 0,
            'failed': 0,
            'success': 0,
        },
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.excel_config = ExcelSecurityConfig()
        self.excel_utils = ExcelMetadataUtils(self.excel_config)

    def get_uploader(self, path, file_specification, organized_files, params: Dict = None):
        """Get uploader from entrypoint."""
        return self.entrypoint(
            self.run, path, file_specification, organized_files, extra_params=params.get('extra_params')
        )

    def _discover_files_recursive(self, dir_path: Path) -> List[Path]:
        """Discover files recursively in a directory."""
        return [file_path for file_path in dir_path.rglob('*') if file_path.is_file()]

    def _discover_files_non_recursive(self, dir_path: Path) -> List[Path]:
        """Discover files in a directory (non-recursive)."""
        return [file_path for file_path in dir_path.glob('*') if file_path.is_file()]

    def _validate_excel_security(self, excel_path: Path) -> None:
        """Validate Excel file security constraints.

        Performs comprehensive security validation including:
        - File size limits
        - Memory usage constraints
        - Basic malicious content detection

        Args:
            excel_path: Path to the Excel file to validate

        Raises:
            ExcelSecurityError: If security validation fails
        """
        # File size check (already done in validator, but double-check)
        file_size = excel_path.stat().st_size
        if file_size > self.excel_config.MAX_FILE_SIZE_BYTES:
            raise ExcelSecurityError(
                f'Excel file too large: {file_size} bytes (max: {self.excel_config.MAX_FILE_SIZE_BYTES})'
            )

        # Memory usage estimation (rough estimate: file_size * 3 for processing)
        estimated_memory = file_size * 3
        if estimated_memory > self.excel_config.MAX_MEMORY_USAGE_BYTES:
            raise ExcelSecurityError(
                f'Excel file may consume too much memory: ~{estimated_memory} bytes '
                f'(max: {self.excel_config.MAX_MEMORY_USAGE_BYTES})'
            )

    def _prepare_excel_file(self, excel_path: Path) -> BytesIO:
        """Prepare Excel file for reading with security validation.

        Args:
            excel_path: Path to the Excel file

        Returns:
            BytesIO: Excel file stream ready for reading

        Raises:
            ExcelSecurityError: If security validation fails
        """
        self._validate_excel_security(excel_path)
        excel_bytes = excel_path.read_bytes()
        return BytesIO(excel_bytes)

    def _process_excel_headers(self, headers: tuple) -> tuple:
        """Process and validate Excel headers.

        Args:
            headers: Raw header tuple from Excel

        Returns:
            tuple: Validated headers

        Raises:
            ExcelParsingError: If headers are invalid
        """
        if len(headers) < 2:
            raise ExcelParsingError('Excel file must have at least 2 columns (file name and metadata)')
        self._validate_excel_content(headers, 0)  # Validate column count
        return headers

    def _process_excel_data_row(self, row: tuple, headers: tuple) -> Optional[Dict[str, Any]]:
        """Process a single Excel data row.

        Args:
            row: Raw row data from Excel
            headers: Excel headers

        Returns:
            Optional[Dict[str, Any]]: Processed row data or None if row should be skipped
        """
        # Skip empty rows
        if not row[0] or str(row[0]).strip() == '':
            return None

        file_name = str(row[0]).strip()
        if not self.excel_utils.is_valid_filename_length(file_name):
            self.run.log_message_with_code('FILENAME_TOO_LONG', file_name[:50])
            return None

        # Create metadata dictionary from remaining columns
        file_metadata: Dict[str, Any] = {}
        for i, value in enumerate(row[1:], start=1):
            if value is not None and i < len(headers):
                header_value = headers[i]
                column_name = str(header_value).strip() if header_value is not None else f'column_{i}'

                # Validate and truncate column name and value
                column_name = self.excel_utils.validate_and_truncate_string(
                    column_name, self.excel_config.MAX_COLUMN_NAME_LENGTH
                )
                str_value = self.excel_utils.validate_and_truncate_string(
                    str(value), self.excel_config.MAX_METADATA_VALUE_LENGTH
                )
                file_metadata[column_name] = str_value

        return {file_name: file_metadata} if file_metadata else None

    def _process_excel_worksheet(self, worksheet) -> Dict[str, Dict[str, Any]]:
        """Process Excel worksheet and extract metadata.

        Args:
            worksheet: openpyxl worksheet object

        Returns:
            Dict[str, Dict[str, Any]]: Extracted metadata dictionary

        Raises:
            ExcelParsingError: If worksheet processing fails
        """
        if worksheet is None:
            raise ExcelParsingError('Excel file has no active worksheet')

        metadata_dict: Dict[str, Dict[str, Any]] = {}
        headers: Optional[tuple] = None
        data_row_count = 0
        validation_interval = getattr(self.excel_config, 'VALIDATION_CHECK_INTERVAL', 1000)

        # Process rows one by one for memory efficiency
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip completely empty rows

            if row_idx == 0:  # Header row
                headers = self._process_excel_headers(row)
                continue

            # Data rows
            if headers is None:
                raise ExcelParsingError('Excel file missing header row')

            data_row_count += 1

            # Validate row count periodically
            if data_row_count % validation_interval == 0:
                self._validate_excel_content(headers, data_row_count)

            # Process individual row
            row_result = self._process_excel_data_row(row, headers)
            if row_result:
                metadata_dict.update(row_result)

        # Final validation
        self._validate_excel_content(headers or (), data_row_count)

        return metadata_dict

    def _validate_excel_content(self, headers: tuple, row_count: int) -> None:
        """Validate Excel content constraints.

        Args:
            headers: Tuple of header values from the first row
            row_count: Total number of data rows processed

        Raises:
            ExcelParsingError: If content validation fails
        """
        # Limit number of columns to prevent memory exhaustion
        if len(headers) > self.excel_config.MAX_COLUMNS:
            raise ExcelParsingError(f'Too many columns: {len(headers)} (max: {self.excel_config.MAX_COLUMNS})')

        # Limit number of rows to prevent excessive processing
        if row_count > self.excel_config.MAX_ROWS:
            raise ExcelParsingError(f'Too many rows: {row_count} (max: {self.excel_config.MAX_ROWS})')

    def _find_excel_metadata_file(self, pathlib_cwd: Path) -> Optional[Path]:
        """Find Excel metadata file in the directory.

        Checks for meta.xlsx and meta.xls in the given directory.

        Args:
            pathlib_cwd (Path): The pathlib object representing the current working directory.

        Returns:
            Optional[Path]: Path to the Excel metadata file if found, None otherwise.
        """
        # Check for xlsx first, then xls
        for extension in ['.xlsx', '.xls']:
            excel_path = pathlib_cwd / f'meta{extension}'
            if excel_path.exists() and excel_path.is_file():
                return excel_path
        return None

    def _read_excel_metadata(self, pathlib_cwd: Path) -> Dict[str, Dict[str, Any]]:
        """Read metadata from excel file with comprehensive security validation.

        This method orchestrates the Excel metadata reading process by delegating
        to specialized methods for each step.

        Args:
            pathlib_cwd (Path): The pathlib object representing the current working directory.

        Returns:
            Dict[str, Dict[str, Any]]: Dictionary mapping file names to their metadata key-value pairs.
            Empty dict if no metadata is configured or if reading fails.

        Raises:
            ExcelSecurityError: If security validation fails
            ExcelParsingError: If Excel content is invalid or exceeds limits
        """
        excel_path = None

        # Check if user provided a specific excel_metadata_path
        excel_metadata_path = self.params.get('excel_metadata_path')
        if excel_metadata_path:
            excel_path = pathlib_cwd / excel_metadata_path
            if not excel_path.exists():
                self.run.log_message_with_code('EXCEL_FILE_NOT_FOUND_PATH')
                return {}
        else:
            # Look for default meta.xlsx or meta.xls
            excel_path = self._find_excel_metadata_file(pathlib_cwd)
            if not excel_path:
                # No Excel metadata file found, return empty dict (not an error)
                return {}

        try:
            self.run.log_message_with_code('EXCEL_FILE_VALIDATION_STARTED')

            # Prepare Excel file with security validation
            excel_stream = self._prepare_excel_file(excel_path)

            # Load and process workbook
            workbook = load_workbook(excel_stream, read_only=True, data_only=True)
            try:
                self.run.log_message_with_code('EXCEL_WORKBOOK_LOADED')
                return self._process_excel_worksheet(workbook.active)
            finally:
                workbook.close()

        except ExcelSecurityError as e:
            self.run.log_message_with_code('EXCEL_SECURITY_VALIDATION_FAILED', str(e))
            raise
        except ExcelParsingError as e:
            self.run.log_message_with_code('EXCEL_PARSING_FAILED', str(e))
            raise
        except InvalidFileException as e:
            self.run.log_message_with_code('EXCEL_INVALID_FILE_FORMAT', str(e))
            raise ExcelParsingError(f'Invalid Excel file format: {str(e)}')
        except MemoryError:
            self.run.log_message_with_code('EXCEL_FILE_TOO_LARGE')
            raise ExcelSecurityError('Excel file exceeds memory limits')
        except (OSError, IOError) as e:
            self.run.log_message_with_code('EXCEL_FILE_ACCESS_ERROR', str(e))
            raise ExcelParsingError(f'File access error: {str(e)}')
        except Exception as e:
            self.run.log_message_with_code('EXCEL_UNEXPECTED_ERROR', str(e))
            raise ExcelParsingError(f'Unexpected error: {str(e)}')

    def start(self) -> Dict[str, Any]:
        """Start upload process.

        Returns:
            Dict: The result of the upload process.
        """
        # Setup result dict early for error handling
        result: Dict[str, Any] = {}

        # Setup path object with path and storage.
        storage_id = self.params.get('storage')
        if storage_id is None:
            raise ActionError('Storage parameter is required')
        storage = self.client.get_storage(storage_id)

        path = self.params.get('path')
        if path is None:
            raise ActionError('Path parameter is required')
        pathlib_cwd = get_pathlib(storage, path)

        # Read excel metadata if configured or default file exists
        excel_metadata: Dict[str, Dict[str, Any]] = {}
        try:
            excel_metadata = self._read_excel_metadata(pathlib_cwd)
            if excel_metadata:
                self.run.log_message_with_code('EXCEL_METADATA_LOADED', len(excel_metadata))
        except ExcelSecurityError as e:
            # Security violations should stop the process entirely
            self.run.log_message_with_code('EXCEL_SECURITY_VIOLATION', str(e))
            return result
        except ExcelParsingError as e:
            # Parsing errors can be non-critical if user didn't explicitly provide Excel file
            if self.params.get('excel_metadata_path'):
                # User explicitly provided Excel file, treat as error
                self.run.log_message_with_code('EXCEL_PARSING_ERROR', str(e))
                return result
            else:
                # Default Excel file found but failed to parse, treat as warning and continue
                self.run.log_message_with_code('EXCEL_PARSING_ERROR', str(e))
                excel_metadata = {}

        # Analyze Collection file specifications to determine the data structure for upload.
        file_specification_template = self._analyze_collection()
        organized_files = self._organize_files(pathlib_cwd, file_specification_template, excel_metadata)

        # Initialize uploader.
        uploader = self.get_uploader(pathlib_cwd, file_specification_template, organized_files, self.params)

        # Get organized files from the uploader (plugin developer's custom implementation)
        # or use the default organization method if uploader doesn't provide valid files
        organized_files = uploader.handle_upload_files()

        # Validate the organized files
        if not self._validate_organized_files(organized_files, file_specification_template):
            self.run.log_message_with_code('VALIDATION_FAILED')
            raise ActionError('Upload is aborted due to validation errors.')

        # Upload files to synapse-backend.
        if not organized_files:
            self.run.log_message_with_code('NO_FILES_FOUND')
            raise ActionError('Upload is aborted due to missing files.')
        # Choose upload method based on async parameter
        if self.params.get('use_async_upload', True):
            uploaded_files = self.run_async(self._upload_files_async(organized_files, 10))
        else:
            uploaded_files = self._upload_files(organized_files)
        result['uploaded_files_count'] = len(uploaded_files)

        # Generate data units for the uploaded data.
        if not uploaded_files:
            self.run.log_message_with_code('NO_FILES_UPLOADED')
            raise ActionError('Upload is aborted due to no uploaded files.')
        generated_data_units = self._generate_data_units(
            uploaded_files, self.params.get('creating_data_unit_batch_size', 1)
        )
        result['generated_data_units_count'] = len(generated_data_units)

        # Setup task with uploaded synapse-backend data units.
        if not generated_data_units:
            self.run.log_message_with_code('NO_DATA_UNITS_GENERATED')
            raise ActionError('Upload is aborted due to no generated data units.')

        # Clean up if temp dir exists
        self._cleanup_temp_directory()

        self.run.log_message_with_code('IMPORT_COMPLETED')
        return result

    def _analyze_collection(self) -> Dict[str, Any]:
        """Analyze Synapse Collection Specifications.

        Returns:
            Dict: The file specifications of the collection.
        """

        # Initialize progress
        self.run.set_progress(0, 2, category='analyze_collection')

        collection_id = self.params.get('data_collection')
        if collection_id is None:
            raise ActionError('Data collection parameter is required')
        self.run.set_progress(1, 2, category='analyze_collection')

        collection = self.run.client.get_data_collection(collection_id)
        self.run.set_progress(2, 2, category='analyze_collection')

        return collection['file_specifications']

    def _upload_files(self, organized_files: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Upload files to synapse-backend.

        Returns:
            Dict: The result of the upload.
        """
        # Initialize progress
        organized_files_count = len(organized_files)
        self.run.set_progress(0, organized_files_count, category='upload_data_files')
        self.run.log_message_with_code('UPLOADING_DATA_FILES')

        client = self.run.client
        collection_id = self.params.get('data_collection')
        if collection_id is None:
            raise ActionError('Data collection parameter is required')
        upload_result = []
        current_progress = 0
        success_count = 0
        failed_count = 0

        # Initialize metrics
        self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')

        for organized_file in organized_files:
            try:
                # Determine if chunked upload should be used based on file size
                use_chunked_upload = self._requires_chunked_upload(organized_file)
                uploaded_data_file = client.upload_data_file(organized_file, collection_id, use_chunked_upload)
                self.run.log_data_file(organized_file, UploadStatus.SUCCESS)
                success_count += 1
                upload_result.append(uploaded_data_file)
            except Exception as e:
                self.run.log_data_file(organized_file, UploadStatus.FAILED)
                self.run.log_message_with_code('FILE_UPLOAD_FAILED', str(e))
                failed_count += 1

            current_progress += 1
            self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')
            self.run.set_progress(current_progress, organized_files_count, category='upload_data_files')

        # Finish progress
        self.run.set_progress(organized_files_count, organized_files_count, category='upload_data_files')

        return upload_result

    def run_async(self, coro: Awaitable[T]) -> T:
        """Run async coroutine safely using asyncio.run().

        This method properly manages event loop lifecycle and prevents
        resource exhaustion from repeated event loop creation.

        Args:
            coro: The coroutine to execute

        Returns:
            The result of the coroutine execution

        Raises:
            RuntimeError: If called from within an existing event loop
        """
        import concurrent.futures

        def _run_in_thread():
            """Run the coroutine in a separate thread to avoid event loop conflicts."""
            return asyncio.run(coro)

        # Check if we're already in an event loop
        try:
            # If this doesn't raise, we're in an event loop
            asyncio.get_running_loop()
            # Run in thread pool to avoid "RuntimeError: cannot be called from a running event loop"
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_run_in_thread)
                return future.result()
        except RuntimeError:
            # No event loop running, safe to use asyncio.run directly
            return asyncio.run(coro)

    async def _upload_files_async(
        self, organized_files: List[Dict[str, Any]], max_concurrent: int = 10
    ) -> List[Dict[str, Any]]:
        """Upload files to synapse-backend asynchronously with concurrency control."""
        # Initialize progress
        organized_files_count = len(organized_files)
        self.run.set_progress(0, organized_files_count, category='upload_data_files')
        self.run.log_message_with_code('UPLOADING_DATA_FILES')

        client = self.run.client
        collection_id = self.params.get('data_collection')
        if collection_id is None:
            raise ActionError('Data collection parameter is required')
        upload_result = []
        success_count = 0
        failed_count = 0

        # Initialize metrics
        self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')

        # Control concurrency with semaphore
        semaphore = asyncio.Semaphore(max_concurrent)

        async def upload_single_file(organized_file):
            async with semaphore:
                loop = asyncio.get_event_loop()
                try:
                    # Determine if chunked upload should be used based on file size
                    use_chunked_upload = self._requires_chunked_upload(organized_file)
                    # Run sync upload_data_file in thread pool
                    uploaded_data_file = await loop.run_in_executor(
                        None, lambda: client.upload_data_file(organized_file, collection_id, use_chunked_upload)
                    )
                    self.run.log_data_file(organized_file, UploadStatus.SUCCESS)
                    return {'status': 'success', 'result': uploaded_data_file}
                except ClientError as e:
                    # Handle API client errors (network, authentication, server errors)
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code('FILE_UPLOAD_FAILED', f'Client error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'client_error', 'retryable': True}
                except (OSError, IOError) as e:
                    # Handle file system errors (file not found, permissions, disk full)
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code('FILE_UPLOAD_FAILED', f'File system error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'file_error', 'retryable': False}
                except MemoryError as e:
                    # Handle out of memory errors (large files)
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code('FILE_UPLOAD_FAILED', f'Memory error (file too large): {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'memory_error', 'retryable': False}
                except asyncio.TimeoutError as e:
                    # Handle timeout errors (slow network, large files)
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code('FILE_UPLOAD_FAILED', f'Upload timeout: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'timeout_error', 'retryable': True}
                except ValueError as e:
                    # Handle data validation errors (invalid file format, metadata issues)
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code('FILE_UPLOAD_FAILED', f'Data validation error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'validation_error', 'retryable': False}
                except Exception as e:
                    # Handle any remaining unexpected errors
                    self.run.log_data_file(organized_file, UploadStatus.FAILED)
                    self.run.log_message_with_code('FILE_UPLOAD_FAILED', f'Unexpected error: {str(e)}')
                    return {'status': 'failed', 'error': str(e), 'error_type': 'unknown_error', 'retryable': False}

        # Create tasks for all files
        tasks = [upload_single_file(organized_file) for organized_file in organized_files]

        # Process files with progress updates
        current_progress = 0
        for completed_task in asyncio.as_completed(tasks):
            result = await completed_task
            current_progress += 1

            if result['status'] == 'success':
                success_count += 1
                upload_result.append(result['result'])
            else:
                failed_count += 1

            # Update metrics and progress
            self._update_metrics(organized_files_count, success_count, failed_count, 'data_files')
            self.run.set_progress(current_progress, organized_files_count, category='upload_data_files')

        # Finish progress
        self.run.set_progress(organized_files_count, organized_files_count, category='upload_data_files')

        return upload_result

    def _generate_data_units(self, uploaded_files: List[Dict[str, Any]], batch_size: int) -> List[Dict[str, Any]]:
        """Generate data units for the uploaded data.

        TODO: make dynamic batch size depend on uploaded file sizes

        Returns:
            Dict: The result of the generate data units process.
        """
        # Initialize progress
        upload_result_count = len(uploaded_files)
        self.run.set_progress(0, upload_result_count, category='generate_data_units')
        self.run.log_message_with_code('GENERATING_DATA_UNITS')

        client = self.run.client
        generated_data_units = []
        current_progress = 0
        success_count = 0
        failed_count = 0

        batches = get_batched_list(uploaded_files, batch_size)
        batches_count = len(batches)

        # Initialize metrics
        self._update_metrics(upload_result_count, success_count, failed_count, 'data_units')

        for batch in batches:
            try:
                created_data_units = client.create_data_units(batch)
                success_count += len(created_data_units)
                generated_data_units.append(created_data_units)
                for created_data_unit in created_data_units:
                    self.run.log_data_unit(
                        created_data_unit['id'], UploadStatus.SUCCESS, data_unit_meta=created_data_unit.get('meta')
                    )
            except Exception as e:
                failed_count += len(batch)
                self.run.log_message_with_code('DATA_UNIT_BATCH_FAILED', str(e))
                for _ in batch:
                    self.run.log_data_unit(None, UploadStatus.FAILED, data_unit_meta=None)

            current_progress += 1
            self._update_metrics(upload_result_count, success_count, failed_count, 'data_units')
            self.run.set_progress(current_progress, batches_count, category='generate_data_units')

        # Finish progress
        self.run.set_progress(upload_result_count, upload_result_count, category='generate_data_units')

        return sum(generated_data_units, [])

    def _validate_organized_files(
        self, organized_files: List[Dict[str, Any]], file_specification_template: Dict[str, Any]
    ) -> bool:
        """Validate organized files from Uploader."""
        validator = FileSpecificationValidator(file_specification_template, organized_files)
        return validator.validate()

    def _organize_files(
        self,
        directory: Path,
        file_specification: List[Dict[str, Any]],
        excel_metadata: Optional[Dict[str, Dict[str, Any]]] = None,
    ) -> List[Dict[str, Any]]:
        """Organize files according to the file specification.

        Args:
            directory (Path): Root directory containing files to organize.
            file_specification (List[Dict[str, Any]]): File specification list with metadata.
            excel_metadata (Optional[Dict[str, Dict[str, Any]]]): Dictionary mapping file names
                to their metadata key-value pairs from excel file.

        Returns:
            List[Dict[str, Any]]: List of dictionaries containing organized files with metadata.
        """
        organized_files: List[Dict[str, Any]] = []

        # Check for type-based directory structure (e.g., image_1/, pcd_1/)
        type_dirs: Dict[str, Path] = {}

        for spec in file_specification:
            spec_name = spec['name']
            spec_dir = directory / spec_name
            if spec_dir.exists() and spec_dir.is_dir():
                type_dirs[spec_name] = spec_dir

        if type_dirs:
            self.run.log_message_with_code('TYPE_DIRECTORIES_FOUND', list(type_dirs.keys()))

        # If type-based directories don't exist, exit early
        if not type_dirs:
            self.run.log_message_with_code('NO_TYPE_DIRECTORIES')
            return organized_files

        self.run.log_message_with_code('TYPE_STRUCTURE_DETECTED')
        self.run.log_message_with_code('FILE_ORGANIZATION_STARTED')

        # Collect and process files in a single pass
        dataset_files = {}
        required_specs = [spec['name'] for spec in file_specification if spec.get('is_required', False)]

        # Get recursive setting from params
        is_recursive = self.params.get('is_recursive', True)

        # Process all files from all type directories
        for spec_name, dir_path in type_dirs.items():
            # Use appropriate method based on recursive setting
            if is_recursive:
                files_list = self._discover_files_recursive(dir_path)
            else:
                files_list = self._discover_files_non_recursive(dir_path)

            for file_path in files_list:
                # Always use filename only for matching
                file_name = file_path.stem

                # Initialize dataset entry if it doesn't exist
                if file_name not in dataset_files:
                    dataset_files[file_name] = {}

                # Map this file to its specification (handle duplicates)
                if spec_name not in dataset_files[file_name]:
                    dataset_files[file_name][spec_name] = file_path
                else:
                    existing_file = dataset_files[file_name][spec_name]
                    if file_path.stat().st_mtime > existing_file.stat().st_mtime:
                        dataset_files[file_name][spec_name] = file_path

        if not dataset_files:
            self.run.log_message_with_code('NO_FILES_FOUND_WARNING')
            return organized_files

        self.run.log_message_with_code('FILES_DISCOVERED', len(dataset_files))

        # Organize datasets - check requirements and create metadata
        for file_name, files_dict in sorted(dataset_files.items()):
            if all(req in files_dict for req in required_specs):
                # Get most common file extension
                file_extensions = {}
                for file_path in files_dict.values():
                    ext = file_path.suffix.lower()
                    if ext:
                        file_extensions[ext] = file_extensions.get(ext, 0) + 1

                origin_file_extension = max(file_extensions.items(), key=lambda x: x[1])[0] if file_extensions else ''

                # Create metadata for this dataset
                meta_data: Dict[str, Any] = {
                    'origin_file_stem': file_name,
                    'origin_file_extension': origin_file_extension,
                    'created_at': datetime.now().isoformat(),
                }

                # Add excel metadata if available
                if excel_metadata and file_name in excel_metadata:
                    meta_data.update(excel_metadata[file_name])

                # Add the organized dataset
                organized_files.append({'files': files_dict, 'meta': meta_data})
            else:
                missing = [req for req in required_specs if req not in files_dict]
                self.run.log_message_with_code('MISSING_REQUIRED_FILES', file_name, ', '.join(missing))

        return organized_files

    def _get_file_size_mb(self, file_path: Path) -> float:
        """Get file size in MB.

        Args:
            file_path (Path): Path to the file.

        Returns:
            float: File size in MB.
        """
        return file_path.stat().st_size / (1024 * 1024)

    def _requires_chunked_upload(self, organized_file: Dict[str, Any]) -> bool:
        """Determine if chunked upload is required based on file size threshold.

        Args:
            organized_file (Dict[str, Any]): Organized file data with 'files' dict.

        Returns:
            bool: True if any file exceeds the threshold, False otherwise.
        """
        max_file_size_mb = self.params.get('max_file_size_mb', 50)
        for file_path in organized_file.get('files', {}).values():
            if isinstance(file_path, Path) and self._get_file_size_mb(file_path) > max_file_size_mb:
                return True
        return False

    def _cleanup_temp_directory(self, temp_path: Optional[Path] = None) -> None:
        """Clean up temporary directory.

        Args:
            temp_path (Optional[Path]): Path to temporary directory.
                If None, uses default temp directory in current working directory.
        """
        if temp_path is None:
            try:
                temp_path = Path(os.getcwd()) / 'temp'
            except (FileNotFoundError, OSError):
                return

        if not temp_path.exists():
            return

        shutil.rmtree(temp_path, ignore_errors=True)
        self.run.log_message(f'Cleaned up temporary directory: {temp_path}')

    def _update_metrics(self, total_count: int, success_count: int, failed_count: int, category: str):
        """Update metrics for upload progress.

        Args:
            total_count (int): Total number of items to process.
            success_count (int): Number of successfully processed items.
            failed_count (int): Number of failed items.
            category (str): The category of the metrics.
        """
        if not self.run:
            raise ValueError('Run instance not properly initialized')

        # Type assertion to help the linter
        assert isinstance(self.run, UploadRun)

        metrics = self.run.MetricsRecord(
            stand_by=total_count - success_count - failed_count, failed=failed_count, success=success_count
        )
        self.run.log_metrics(metrics, category)
