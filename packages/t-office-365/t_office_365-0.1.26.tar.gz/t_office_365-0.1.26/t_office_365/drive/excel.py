"""Excel class."""
from typing import List, Tuple

import requests
from O365 import Account
from openpyxl.utils import get_column_letter

from t_office_365.core import Core
from t_office_365.decorators import retry_if_exception
from t_office_365.endpoints import workbook_api
from t_office_365.models import Cell
from t_office_365.utils import check_result, json_to_cells, json_to_row_cells


class Excel(Core):
    """Excel class is used for API calls to Excel."""

    def __init__(self, account: Account, drive_id: str) -> None:
        """Initializes an instance of the Excel class.

        :param:
        - account (O365.Account): The O365 Account used for authentication.
        - drive_id (str): The ID of the drive.
        """
        super().__init__(account)
        self.__drive_id = drive_id

    @retry_if_exception
    def get_sheet_names(self, file_id: str) -> dict:
        """Get sheet names from Excel file.

        :param:
        - file_id (str): The ID of the Excel file.

        :return:
            dict: The sheet names json.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(workbook_api.worksheets(self.__drive_id, file_id))
        result = requests.get(url, headers=self.headers())
        check_result(result, f"{file_id}")
        return result.json()

    @retry_if_exception
    def create_sheet(self, file_id: str, sheet_name: str) -> None:
        """Create a new sheet in Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the new sheet.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(workbook_api.worksheets(self.__drive_id, file_id))
        payload = {"name": sheet_name}
        result = requests.post(url, json=payload, headers=self.headers())
        check_result(result, f"{file_id}")

    @retry_if_exception
    def get_rows_values(
        self, file_id: str, sheet_name: str, end_row: int = None, end_column: int = None
    ) -> List[List[str]]:
        """Get rows values from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - end_row (int): The index of the last row to retrieve.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        result = self.get_rows_range_data(file_id, sheet_name, end_row=end_row, end_column=end_column)
        return result["values"]

    @retry_if_exception
    def get_rows_cells(
        self, file_id: str, sheet_name: str, end_row: int = None, end_column: int = None
    ) -> List[List[Cell]]:
        """Get rows values from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - end_row (int): The index of the last row to retrieve.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        result = self.get_rows_range_data(file_id, sheet_name, end_row=end_row, end_column=end_column)
        return json_to_cells(result)

    @retry_if_exception
    def get_row_values(self, file_id: str, sheet_name: str, row: int, end_column: int = None) -> List[str]:
        """Get rows values from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - row (int): The row index of the sheet to retrieve.
        - end_column (str): The max column index to retrieve.
        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        result = self.get_rows_range_data(file_id, sheet_name, start_row=row, end_row=row, end_column=end_column)
        try:
            return result["values"][0]
        except IndexError:
            raise ValueError(f"Unable to read values in {row} row.")

    @retry_if_exception
    def get_row_cells(
        self, file_id: str, sheet_name: str, row_number: int, headers_row: int = 1, end_column: int = None
    ) -> List[Cell]:
        """Get rows values from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - row (int): The row index of the sheet to retrieve.
        - end_column (str): The max column index to retrieve.
        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        headers = self.get_rows_range_data(
            file_id, sheet_name, start_row=headers_row, end_row=headers_row, end_column=end_column
        )
        result = self.get_rows_range_data(
            file_id, sheet_name, start_row=row_number, end_row=row_number, end_column=end_column
        )
        try:
            return json_to_row_cells(result, headers, row_number)
        except IndexError:
            raise ValueError(f"Unable to read values in {row_number} row.")

    @retry_if_exception
    def get_rows_range_data(
        self,
        file_id: str,
        sheet_name: str,
        start_row: int = 1,
        end_row: int = None,
        start_column: int = 1,
        end_column: int = None,
    ) -> dict:
        """Get rows range data from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - end_row (int, optional): The index of the last row to retrieve.
        - end_column (int, optional): The index of the column to retrieve.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        last_row, last_column = self.get_last_row_column(file_id, sheet_name)
        if not end_row:
            end_row = last_row
        if not end_column:
            end_column = last_column
        url = self.get_url(
            workbook_api.get_row_endpoint(
                drive_id=self.__drive_id,
                file_id=file_id,
                sheet_name=sheet_name,
                start_row=start_row,
                end_row=end_row,
                start_column=get_column_letter(start_column),
                end_column=get_column_letter(end_column),
            )
        )
        result = requests.get(url, headers=self.headers())
        check_result(result, f"{file_id}")
        return result.json()

    @retry_if_exception
    def get_cell_value(self, file_id: str, sheet_name: str, row: int, column: int) -> str:
        """Get cell value from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - row (int): The row index of the row.
        - column (int): The column index of the cell.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(workbook_api.get_cell_endpoint(self.__drive_id, file_id, sheet_name, row - 1, column - 1))
        result = requests.get(url, headers=self.headers())
        check_result(result, f"Getting cell value from file {file_id}")
        try:
            return result.json()["values"][0][0]
        except IndexError:
            raise ValueError(f"Unable to read cell in {row} row, {column} column.")

    @retry_if_exception
    def get_cell(self, file_id: str, sheet_name: str, row: int, column: int) -> Cell:
        """Get cell from Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - row (int): The row index of the row.
        - column (int): The column index of the cell.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(workbook_api.get_cell_endpoint(self.__drive_id, file_id, sheet_name, row - 1, column - 1))
        result = requests.get(url, headers=self.headers())
        check_result(result, f"Getting cell value from file {file_id}")
        return Cell(
            "",
            result.json().get("numberFormat", [])[0][0],
            result.json().get("values", None)[0][0],
            result.json().get("rowIndex", 0),
            result.json().get("columnIndex", 0),
        )

    @retry_if_exception
    def update_cell_value(self, file_id: str, sheet_name: str, row: int, column: int, value: str) -> None:
        """Update cell value in Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - row (int): The row index of the sheet to update.
        - column (int): The column index to update.
        - value: The new value to be set in the cell.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(workbook_api.get_cell_endpoint(self.__drive_id, file_id, sheet_name, row - 1, column - 1))
        payload = {"values": [[value]]}
        result = requests.patch(url, json=payload, headers=self.headers())
        check_result(result, f"{file_id}")

    @retry_if_exception
    def update_row_values(self, file_id: str, sheet_name: str, values: List[str], row: int) -> None:
        """Update row values in Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - values: The new values to be set in the row.
        - row (int): The row index of the sheet to update.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(
            workbook_api.get_row_endpoint(
                drive_id=self.__drive_id,
                file_id=file_id,
                sheet_name=sheet_name,
                start_row=row,
                end_row=row,
                end_column=get_column_letter(len(values)),
            )
        )
        payload = {"values": [values]}
        result = requests.patch(url, json=payload, headers=self.headers())
        check_result(result, f"{file_id}")

    @retry_if_exception
    def update_rows_value(
        self, file_id: str, sheet_name: str, values: List[list[str]], start_row: int = 1, end_row: int = 1
    ) -> None:
        """Update rows in Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.
        - values: The new values to be set in the row.
        - start_row (int): The start row index of the sheet to update.
        - end_row (int): The end row index of the sheet to update.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        url = self.get_url(
            workbook_api.get_row_endpoint(
                drive_id=self.__drive_id,
                file_id=file_id,
                sheet_name=sheet_name,
                start_row=start_row,
                end_row=end_row,
                end_column=get_column_letter(len(values[0])),
            )
        )
        payload = {"values": values}
        result = requests.patch(url, json=payload, headers=self.headers())
        check_result(result, f"{file_id}")

    def get_last_row_number(self, file_id: str, sheet_name: str) -> int:
        """Get the last row number in Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        return self.get_last_row_column(file_id, sheet_name)[0]

    def get_last_column_number(self, file_id: str, sheet_name: str) -> int:
        """Get the last column number in Excel.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        return self.get_last_row_column(file_id, sheet_name)[1]

    @retry_if_exception
    def get_last_row_column(self, file_id: str, sheet_name: str) -> Tuple[int, int]:
        """Get File ID in Excel by excel path in site.

        :param:
        - file_id (str): The ID of the Excel file.
        - sheet_name (str): The name of the sheet.

        :raises:
        - BadRequestError: If there is a bad request.
        - UnexpectedError: If there is an unexpected error during the request.
        """
        result = requests.get(
            self.get_url(workbook_api.get_used_range(self.__drive_id, file_id, sheet_name)), headers=self.headers()
        )
        check_result(result, f"{file_id}")
        range_info = result.json()
        return int(range_info["rowCount"]), int(range_info["columnCount"])
