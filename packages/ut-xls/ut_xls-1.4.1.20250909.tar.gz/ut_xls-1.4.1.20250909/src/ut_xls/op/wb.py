from typing import Any, TypeAlias

import openpyxl as op
import pandas as pd

from ut_aod.aod import AoD
from ut_dic.dic import Dic

from ut_xls.op.doaoa import DoAoA
from ut_xls.op.ws import Ws

TyCe: TypeAlias = op.cell.cell.Cell
TyWb: TypeAlias = op.workbook.workbook.Workbook
TyWs: TypeAlias = op.worksheet.worksheet.Worksheet
TyCs: TypeAlias = op.chartsheet.chartsheet.Chartsheet
TyPdDf: TypeAlias = pd.DataFrame

TyArr = list[Any]
TyAoA = list[TyArr]
TyAoAoA = list[TyAoA]
TyDic = dict[Any, Any]
TyAoD = list[TyDic]
TyAoS = list[str]
TyAoWs = list[TyWs]
TyDoD = dict[Any, TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoWs = dict[Any, TyWs]
TyDoPdDf = dict[Any, TyPdDf]
TyAoD_DoAoD = TyAoD | TyDoAoD
TySheet = int | str
TyOpSheets = TySheet | list[int | str]
TySheetname = str
TySheetnames = list[TySheetname]
TyStrArr = str | TyArr
TyToCe = tuple[TyCe, ...]
TyCsWs = TyWs | TyCs | TyWs

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnAoD_DoAoD = None | TyAoD_DoAoD
TnAoWs = None | TyAoWs
TnDoWs = None | TyDoWs
TnOpSheets = None | TyOpSheets
TnSheet = None | TySheet
TnSheetname = None | TySheetname
TnWb = None | TyWb
TnWs = None | TyWs
TnCs = None | TyCs
TnCsWs = None | TyCsWs


class Wb:

    @staticmethod
    def iter_sheet_names(wb: TyWb, **kwargs):
        cols_count = kwargs.get('cols_count', 0)
        sheet_names: TyArr = kwargs.get('sheet_names', [])
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            if sheet.max_column == cols_count:
                yield sheet_name

    @staticmethod
    def iter_sheet(wb: TyWb, max_sheets):
        for _ii in range(0, max_sheets):
            yield wb.create_sheet()

    @staticmethod
    def sh_sheetname_in_arr(sheet: TnSheet, sheetnames: TySheetnames) -> TnSheetname:
        if sheet is None:
            return None
        if isinstance(sheet, int):
            if sheet < len(sheetnames):
                return sheetnames[sheet]
            return None
        if isinstance(sheet, str):
            if sheet in sheetnames:
                return sheet
            return None
        return None

    @classmethod
    def sh_sheetname(cls, wb: TnWb, sheet: TnSheet) -> TnSheetname:
        if wb is None or sheet is None:
            return None
        _sheetname: TnSheetname = cls.sh_sheetname_in_arr(sheet, wb.sheetnames)
        return _sheetname

    @classmethod
    def sh_sheetnames(cls, wb: TnWb, sheet: TnOpSheets) -> TySheetnames:
        if wb is None or not sheet:
            _sheetnames: TySheetnames = []
            return _sheetnames
        if isinstance(sheet, (int, str)):
            _sheetname: TnSheetname = cls.sh_sheetname_in_arr(sheet, wb.sheetnames)
            if not _sheetname:
                _sheetnames = []
            else:
                _sheetnames = [_sheetname]
            return _sheetnames
        if isinstance(sheet, (list, tuple)):
            _sheetnames = []
            for _sheet in sheet:
                _sheetname = cls.sh_sheetname_in_arr(_sheet, wb.sheetnames)
                if _sheetname:
                    _sheetnames.append(_sheetname)
            return _sheetnames
        _sheetnames = []
        return _sheetnames

    @classmethod
    def sh_sheet_by_sheetname(
            cls, wb: TnWb, sheetname: TnSheetname) -> TnWs:
        if wb is None:
            return None
        if not sheetname:
            return None
        return wb[sheetname]

    @classmethod
    def sh_sheet(cls, wb: TnWb, sheet: TySheet) -> TnWs:
        _ws: TnWs = cls.sh_sheet_by_sheetname(wb, cls.sh_sheetname(wb, sheet))
        return _ws

    # @classmethod
    # def sh_sheet_by_type(
    #         cls, wb: TnWb, sheet_name: TnSheetname, sheet_type: str
    # ) -> TnCsWs:
    #     _ws: TnCsWs = Ws.sh_by_type(
    #             cls.sh_sheet_by_sheetname(wb, sheet_name), sheet_type)
    #     return _ws

    @classmethod
    def sh_chartsheet_by_sheetname(cls, wb: TnWb, sheet_name: TnSheetname) -> TnCs:
        _cs: TnCs = Ws.sh_chartsheet(cls.sh_sheet_by_sheetname(wb, sheet_name))
        return _cs

    @classmethod
    def sh_worksheet_by_sheetname(cls, wb: TnWb, sheet_name: TnSheetname) -> TnWs:
        _ws: TnWs = Ws.sh_worksheet(cls.sh_sheet_by_sheetname(wb, sheet_name))
        return _ws

    @classmethod
    def sh_chartsheet(cls, wb: TnWb, sheet: TnSheet) -> TnCs:
        return cls.sh_chartsheet_by_sheetname(wb, cls.sh_sheetname(wb, sheet))

    @classmethod
    def sh_worksheet(cls, wb: TnWb, sheet: TnSheet) -> TnWs:
        return cls.sh_worksheet_by_sheetname(wb, cls.sh_sheetname(wb, sheet))

    @classmethod
    def to_aod(cls, wb: TnWb, sheet: TnSheet) -> TyAoD:
        if wb is None:
            return []
        _ws: TnWs = cls.sh_worksheet(wb, sheet)
        _aod: TyAoD = Ws.to_aod(_ws)
        return _aod

    @classmethod
    def to_doaod(cls, wb: TnWb, sheet: TnOpSheets) -> TyDoAoD:
        if wb is None:
            return {}
        doaod: TyDoAoD = {}
        if wb is None:
            return doaod
        _sheetnames: TySheetnames = cls.sh_sheetnames(wb, sheet)
        if not _sheetnames:
            return doaod
        for _sheetname in _sheetnames:
            _ws: TnWs = cls.sh_worksheet_by_sheetname(wb, _sheetname)
            Dic.set_kv_not_none(doaod, _sheetname, Ws.to_aod(_ws))
        return doaod

    @classmethod
    def to_aod_or_doaod(
            cls, wb: TyWb, sheet: TnOpSheets) -> TyAoD_DoAoD:
        doaod: TyDoAoD = {}
        _sheetnames: TySheetnames = cls.sh_sheetnames(wb, sheet)
        if not _sheetnames:
            return doaod
        if len(_sheetnames) == 1:
            _sheetname = _sheetnames[0]
            _ws: TnWs = Wb.sh_worksheet_by_sheetname(wb, _sheetname)
            _aod: TyAoD = Ws.to_aod(_ws)
            return _aod
        for _sheetname in _sheetnames:
            _ws = Wb.sh_worksheet_by_sheetname(wb, _sheetname)
            Dic.set_kv_not_none(doaod, _sheetname, Ws.to_aod(_ws))
        return doaod

    @classmethod
    def createupdate_wb_with_doaoa(cls, wb: TnWb, doaoa: TnDoAoA) -> None:
        if not doaoa:
            return
        if wb is None:
            DoAoA.create_wb(doaoa)
        else:
            cls.update_wb_with_doaoa(wb, doaoa)

    @classmethod
    def update_wb_with_aoa(cls, wb: TnWb, aoa: TnAoA, sheet_name: str) -> None:
        if wb is None:
            return
        if not aoa:
            return
        _sheet_name: TnSheetname = cls.sh_sheetname(wb, sheet_name)
        _ws: TnWs = cls.sh_worksheet_by_sheetname(wb, _sheet_name)
        Ws.append_rows(_ws, aoa)

    @classmethod
    def update_wb_with_aod(cls, wb: TnWb, aod: TnAoD, sheet_name: str) -> None:
        if wb is None:
            return
        _aoa: TnAoA = AoD.to_aoa(aod, sw_keys=False)
        cls.update_wb_with_aoa(wb, _aoa, sheet_name)

    @classmethod
    def update_wb_with_doaoa(cls, wb: TnWb, doaoa: TnDoAoA) -> None:
        if wb is None:
            return
        if not doaoa:
            return
        a_ws_id: TyArr = Dic.sh_keys(doaoa, wb.sheetnames)
        for ws_id in a_ws_id:
            aoa: TyAoA = doaoa[ws_id]
            ws: TnWs = cls.sh_worksheet(wb, ws_id)
            Ws.append_rows(ws, aoa)

    @staticmethod
    def update_wb_with_dodf(wb: TnWb, dodf: TyDoPdDf, **kwargs) -> TnWb:
        if wb is None:
            return None
        _d_update: TyDic = kwargs.get('d_update', {})
        _d_head: TyDic = _d_update.get('d_head', {})
        _a_key: TyArr = Dic.show_sorted_keys(dodf)
        for _key in _a_key:
            _df = dodf[_key]
            _ws_tpl: TyWs = wb['TMPL']
            _ws_new: TyWs = wb.copy_worksheet(_ws_tpl)
            _ws_new.title = _key
            _d_head['title']['value'] = _key
            Ws.update_ws_cell_from_df_with_d_body(_ws_new, _df, _d_update)
            Ws.update_ws_cell_with_d_head(_ws_new, _d_head)
        return wb
