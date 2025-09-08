from typing import Any, TypeAlias, IO, TYPE_CHECKING

from os import PathLike
import openpyxl as op
# from pathlib import Path

from ut_obj.io import Io

from ut_xls.op.ioipathwb import IoiPathWb
from ut_xls.op.wb import Wb
from ut_xls.op.ws import Ws

if TYPE_CHECKING:
    from _typeshed import SupportsRead
    TyOpFileSrc = str | PathLike[str] | IO[bytes] | SupportsRead[bytes]
else:
    TyOpFileSrc = str | PathLike[str] | IO[bytes]

TyWb: TypeAlias = op.workbook.workbook.Workbook
TyWs: TypeAlias = op.worksheet.worksheet.Worksheet

TyArr = list[Any]
TyDic = dict[Any, Any]
TyAoA = list[TyArr]
TyAoD = list[TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoWs = dict[Any, TyWs]
TyAoD_DoAoD = TyAoD | TyDoAoD
# TyOpFileSrc = str | bytes | Path | TextIO | BinaryIO
TyPath = str
TyPathnm = str

TySheet = int | str
TySheets = int | str | list[int | str]
TySheetname = str
TySheetnames = list[TySheetname]

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnAoD_DoAoD = None | TyAoD_DoAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnDoAoD = None | TyDoAoD
TnDoWs = None | TyDoWs
TnSheet = None | TySheet
TnSheets = None | TySheets
TnSheetname = None | TySheetname
TnSheetnames = None | TySheetnames
TnWs = None | TyWs


class IoiPathWs:

    @staticmethod
    def read_ws_to_dic(
            io: TyOpFileSrc, sheet: TySheet) -> TnDic:
        _dic: TyDic = Ws.to_dic(Wb.sh_sheet(IoiPathWb.load(io), sheet))
        return _dic

    @staticmethod
    def read_ws_to_aod(
            io: TyOpFileSrc, sheet: TySheet) -> TnAoD:
        _aod: TyAoD = Ws.to_aod(Wb.sh_sheet(IoiPathWb.load(io), sheet))
        return _aod

    @staticmethod
    def read_ws_filter_rows(io: TyOpFileSrc, sheet: TySheet) -> TnArr:
        Io.verify(io)
        _arr: TnArr = Ws.filter_rows(Wb.sh_sheet(IoiPathWb.load(io), sheet))
        return _arr

    @staticmethod
    def read_ws_to_aoa(
            io: TyOpFileSrc, sheet: TnSheets = None) -> tuple[TnAoA, TnSheetnames]:
        Io.verify(io)
        _wb: TyWb = IoiPathWb.load(io)
        aoa: TyAoA = []
        if not sheet:
            return aoa, None
        _sheetnames: TnSheetnames = Wb.sh_sheetnames(_wb, sheet)
        if not _sheetnames:
            return aoa, _sheetnames
        for _sheetname in _sheetnames:
            _ws: TnWs = Wb.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                values: TyArr = Ws.to_row_values(_ws)
                aoa.append(values)
        return aoa, _sheetnames

    @staticmethod
    def read_sheetnames(io: TyOpFileSrc) -> TyArr:
        Io.verify(io)
        wb: TyWb = IoiPathWb.load(io)
        sheetnames: TySheetnames = wb.sheetnames
        return sheetnames

    @staticmethod
    def read_ws_to_doaoa(
            io: TyOpFileSrc, sheet: TnSheets = None) -> tuple[TnDoAoA, TnSheetnames]:
        Io.verify(io)
        _wb: TyWb = IoiPathWb.load(io)
        doaoa: TyDoAoA = {}
        if _wb is None:
            return doaoa, None
        sheetnames: TnSheetnames = Wb.sh_sheetnames(_wb, sheet)
        if not sheetnames:
            return doaoa, sheetnames
        for _sheetname in sheetnames:
            _ws: TnWs = Wb.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                values: TyArr = Ws.to_row_values(_ws)
                doaoa[sheet] = values
        return doaoa, sheetnames

    @staticmethod
    def read_ws_to_dowsop(
            io: TyOpFileSrc, sheet: TnSheets = None) -> tuple[TnDoWs, TnSheetnames]:
        Io.verify(io)
        _wb: TyWb = IoiPathWb.load(io)
        dows: TyDoWs = {}
        if _wb is None:
            return dows, None
        sheetnames: TnSheetnames = Wb.sh_sheetnames(_wb, sheet)
        if not sheetnames:
            return dows, sheetnames
        for _sheetname in sheetnames:
            _ws: TnWs = Wb.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                dows[_sheetname] = _ws
        return dows, sheetnames
