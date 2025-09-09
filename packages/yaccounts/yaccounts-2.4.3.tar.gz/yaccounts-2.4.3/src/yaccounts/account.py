import calendar
from dataclasses import dataclass
from datetime import datetime
import enum
import pathlib

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import SheetFormatProperties
import pandas as pd

from .budget_categories import AccountType, BudgetType
from .config import DEBUG_WITH_CACHING, DEVELOPER_NAME, END_MONTH
from .run_report import (
    run_activity_report,
    run_gift_report,
    run_grant_report,
    run_journal_report,
    run_payroll_report,
    run_transactions_report,
)
from .yutils import TermColors, print_color, warn, workday_str_amount_to_float, error

ROW_INDENT_STR = " ▸ "

MONTH_COLS = calendar.month_abbr[1:]


@dataclass(frozen=True)
class _ColumnHeader:
    col_name: str
    index: int


class _ColumnHeaders(enum.Enum):
    """Enum for column indices in the DataFrame."""

    CATEGORY = _ColumnHeader("Category", 1)
    BUDGET = _ColumnHeader("Budget", 2)
    PREV_ACTUALS = _ColumnHeader("Prev Actuals", 3)
    JAN = _ColumnHeader("Jan", 4)
    FEB = _ColumnHeader("Feb", 5)
    MAR = _ColumnHeader("Mar", 6)
    APR = _ColumnHeader("Apr", 7)
    MAY = _ColumnHeader("May", 8)
    JUN = _ColumnHeader("Jun", 9)
    JUL = _ColumnHeader("Jul", 10)
    AUG = _ColumnHeader("Aug", 11)
    SEP = _ColumnHeader("Sep", 12)
    OCT = _ColumnHeader("Oct", 13)
    NOV = _ColumnHeader("Nov", 14)
    DEC = _ColumnHeader("Dec", 15)
    ACTUALS_YTD = _ColumnHeader("YTD", 16)
    COMMITTED = _ColumnHeader("Pending Commitments", 17)
    BALANCE = _ColumnHeader("Balance", 18)


def _get_column_headers():
    """Get the column headers as a list of strings."""
    return [col.value.col_name for col in _ColumnHeaders]


class Account:
    def __init__(
        self,
        account_code,
        year,
        transactions_xlsx_out_path=None,
        payroll_xlsx_out_path=None,
        journals_xlsx_out_path=None,
    ):

        self.account_code = account_code
        self.year = year
        self.transactions_xlsx_out_path = transactions_xlsx_out_path
        self.payroll_xlsx_out_path = payroll_xlsx_out_path
        self.journals_xlsx_out_path = journals_xlsx_out_path

        try:
            self.account_type = AccountType[account_code[:2].upper()]
        except KeyError:
            error(
                f"Account code {self.account_code} does not start with 'GR', 'AC' or 'GF'. "
                "Other account codes are not supported yet."
            )

        if self.account_type == AccountType.GR:
            assert self.payroll_xlsx_out_path is None, (
                "Payroll data is not applicable for GR accounts. "
                "Please set payroll_xlsx_out_path to None."
            )
            assert self.journals_xlsx_out_path is None, (
                "Journal data is not applicable for GR accounts. "
                "Please set journals_xlsx_out_path to None."
            )
        else:
            assert self.transactions_xlsx_out_path is None, (
                "Transaction data is not applicable for non-GR accounts. "
                "Please set transactions_xlsx_out_path to None."
            )

        self._df = pd.DataFrame(columns=_get_column_headers())

        self._df_transactions = None
        self._df_journals = None
        self._df_payroll = None

    def get_workday_data(self, dr):
        """Get all workday data for the account."""

        # This will run the workday reports.
        # First the annual reports are run (transactaions for GR accounts, payroll for non-GR accounts),
        # Then the month-by-month reports are run to get monthly balances (and monthly journal transactions for non-GR accounts).
        # This saves all the workday data into pandas DataFrames (self._df* properties).
        if self.uses_transaction_report_data():
            self._get_transactions(dr)
        else:
            self._get_payroll_data(dr)
        self._get_monthly_data(dr)

        # Next, the data is processed and merged into the main DataFrame (self._df).
        self._finalize_workday_data()
        self._add_transaction_or_journal_data()

        if not self.uses_transaction_report_data():
            self._add_payroll_data()

    def uses_transaction_report_data(self):
        """Check if the account uses data from the new grant transaction report (only available to GR accounts)."""
        return self.account_type == AccountType.GR

    ############################################################################
    ##################### Methods to get data from workday #####################
    ############################################################################

    def _get_transactions(self, dr):
        """Get transactions from Workday."""

        cache_filename = pathlib.Path(
            f"cache/transactions_{self.account_code}_{self.year}.pkl"
        )
        if DEBUG_WITH_CACHING and cache_filename.exists():
            print_color(
                TermColors.BLUE,
                f"Loading transactions from cache: {cache_filename}",
            )
            self._df_transactions = pd.read_pickle(cache_filename)
            return

        self._df_transactions = run_transactions_report(
            dr, self.account_code, self.year
        )
        if DEBUG_WITH_CACHING:
            cache_filename.parent.mkdir(parents=True, exist_ok=True)
            self._df_transactions.to_pickle(cache_filename)
            self._df_transactions.to_csv(
                cache_filename.with_suffix(".csv"), index=False
            )

    def _get_payroll_data(self, dr):
        """Get payroll data from Workday."""

        cache_filename = pathlib.Path(
            f"cache/payroll_{self.account_code}_{self.year}.pkl"
        )
        if DEBUG_WITH_CACHING and cache_filename.exists():
            print_color(
                TermColors.BLUE,
                f"Loading payroll data from cache: {cache_filename}",
            )
            self._df_payroll = pd.read_pickle(cache_filename)
            return

        self._df_payroll = run_payroll_report(dr, self.account_code, self.year)
        if DEBUG_WITH_CACHING:
            cache_filename.parent.mkdir(parents=True, exist_ok=True)
            self._df_payroll.to_pickle(cache_filename)
            self._df_payroll.to_csv(cache_filename.with_suffix(".csv"), index=False)

    def _get_monthly_data(self, dr):
        """Get monthly data from Workday."""

        cache_filename_monthly = pathlib.Path(
            f"cache/monthly_{self.account_code}_{self.year}.pkl"
        )
        cache_filename_journals = pathlib.Path(
            f"cache/journals_{self.account_code}_{self.year}.pkl"
        )

        if (
            DEBUG_WITH_CACHING
            and cache_filename_monthly.exists()
            and cache_filename_journals.exists()
        ):
            print_color(
                TermColors.BLUE,
                f"Loading monthly data from cache: {cache_filename_monthly}",
            )
            self._df = pd.read_pickle(cache_filename_monthly)
            print_color(
                TermColors.BLUE,
                f"Loading journal data from cache: {cache_filename_journals}",
            )
            self._df_journals = pd.read_pickle(cache_filename_journals)
            return

        end_month = 12 if self.year != datetime.now().year else datetime.now().month
        if END_MONTH is not None:
            end_month = END_MONTH
        month_year_list = [
            datetime(self.year, month, 1) for month in range(1, end_month + 1)
        ]

        for month in month_year_list:

            if self.account_code.startswith("GR"):
                data = run_grant_report(dr, self.account_code, month)
            elif self.account_code.startswith("AC"):
                data = run_activity_report(dr, self.account_code, month)
            elif self.account_code.startswith("GF"):
                data = run_gift_report(dr, self.account_code, month)
            if self.account_type != AccountType.GR:
                journal_data_df = run_journal_report(dr, self.account_code, month)

                if self._df_journals is None:
                    self._df_journals = journal_data_df.copy()
                else:
                    self._df_journals = pd.concat(
                        [self._df_journals, journal_data_df], ignore_index=True
                    )

            for category in data:
                # If the actuals_df does not have a row for this category, add it
                if (
                    category
                    not in self._df[_ColumnHeaders.CATEGORY.value.col_name].values
                ):
                    new_row = {_ColumnHeaders.CATEGORY.value.col_name: category}
                    self._df = pd.concat(
                        [self._df, pd.DataFrame([new_row])], ignore_index=True
                    )

                # Update the actuals_df for this category and month
                self._df.loc[
                    self._df[_ColumnHeaders.CATEGORY.value.col_name] == category,
                    month.strftime("%b"),
                ] = data[category][BudgetType.ACTUALS]

                # If January, update previous actuals
                if month.month == 1:
                    self._df.loc[
                        self._df[_ColumnHeaders.CATEGORY.value.col_name] == category,
                        _ColumnHeaders.PREV_ACTUALS.value.col_name,
                    ] = data[category][BudgetType.ACTUALS_PREV]

                # If month is last month of loop, update budget and committed
                if month.month == end_month:
                    self._df.loc[
                        self._df[_ColumnHeaders.CATEGORY.value.col_name] == category,
                        _ColumnHeaders.BUDGET.value.col_name,
                    ] = data[category][BudgetType.BUDGET]
                    self._df.loc[
                        self._df[_ColumnHeaders.CATEGORY.value.col_name] == category,
                        _ColumnHeaders.COMMITTED.value.col_name,
                    ] = data[category][BudgetType.COMMITTED]

        if DEBUG_WITH_CACHING:
            cache_filename_monthly.parent.mkdir(parents=True, exist_ok=True)
            self._df.to_pickle(cache_filename_monthly)
            self._df.to_csv(cache_filename_monthly.with_suffix(".csv"), index=False)
            if self._df_journals is not None:
                self._df_journals.to_pickle(cache_filename_journals)
                self._df_journals.to_csv(
                    cache_filename_journals.with_suffix(".csv"), index=False
                )

    ############################################################################
    ##################### Methods to process workday data ######################
    ############################################################################

    def _finalize_workday_data(self):
        """Finalize and cleanup all workday data"""
        #################### Process actuals (self._df) ####################
        # Convert all columns except 'Category' to float
        for col in self._df.columns:
            if col != _ColumnHeaders.CATEGORY.value.col_name:
                self._df[col] = self._df[col].apply(
                    lambda x: float(x) if x != "" else ""
                )

        # Replace NaN values with 0
        self._df = self._df.fillna(0.0)

        # Sort by Category
        self._df = self._df.sort_values(
            by=_ColumnHeaders.CATEGORY.value.col_name
        ).reset_index(drop=True)

        # YTD column is the sum of all months
        self._df[_ColumnHeaders.ACTUALS_YTD.value.col_name] = self._df[
            [datetime(1900, month, 1).strftime("%b") for month in range(1, 13)]
        ].sum(axis=1)

        # Balance is Budget - Prev Actuals - YTD - Committed
        self._df[_ColumnHeaders.BALANCE.value.col_name] = (
            self._df[_ColumnHeaders.BUDGET.value.col_name]
            - self._df[_ColumnHeaders.PREV_ACTUALS.value.col_name]
            - self._df[_ColumnHeaders.ACTUALS_YTD.value.col_name]
            - self._df[_ColumnHeaders.COMMITTED.value.col_name]
        )

        # Add a totals row
        total_row = self._df.sum(numeric_only=True)
        total_row[_ColumnHeaders.CATEGORY.value.col_name] = "Total"
        total_row[_ColumnHeaders.BUDGET.value.col_name] = total_row[
            _ColumnHeaders.BUDGET.value.col_name
        ]
        total_row[_ColumnHeaders.PREV_ACTUALS.value.col_name] = total_row[
            _ColumnHeaders.PREV_ACTUALS.value.col_name
        ]
        total_row[_ColumnHeaders.BALANCE.value.col_name] = total_row[
            _ColumnHeaders.BALANCE.value.col_name
        ]
        total_row[_ColumnHeaders.ACTUALS_YTD.value.col_name] = total_row[
            _ColumnHeaders.ACTUALS_YTD.value.col_name
        ]
        self._df = pd.concat([self._df, pd.DataFrame([total_row])], ignore_index=True)

        if self.uses_transaction_report_data():
            self._cleanup_transaction_data()
        else:
            self._cleanup_journal_data()

    def _cleanup_transaction_data(self):
        """Cleanup self._df_transactions)"""
        # Remove any rows where 'Accounting Date' is empty (these are total rows and not useful)
        self._df_transactions = self._df_transactions[
            self._df_transactions["Accounting Date"].notna()
            & (self._df_transactions["Accounting Date"].str.strip() != "")
        ].copy()

        # Export transactions to Excel if specified
        if self.transactions_xlsx_out_path:
            self._df_transactions.to_excel(
                self.transactions_xlsx_out_path, index=False, engine="openpyxl"
            )

        # Split 'Ledger Account' on ':' and convert to int
        self._df_transactions["Ledger Account"] = (
            self._df_transactions["Ledger Account"].str.split(":").str[0].astype(int)
        )

    def _cleanup_journal_data(self):
        # Remove any rows where 'Accounting Date' is empty (these are total rows and not useful)
        self._df_journals = self._df_journals[
            self._df_journals["Accounting Date"].notna()
            & (self._df_journals["Accounting Date"].str.strip() != "")
        ].copy()

        # Convert 'Ledger Account by Identifier' to int
        self._df_journals["Ledger Account by Identifier"] = self._df_journals[
            "Ledger Account by Identifier"
        ].astype(int)

    def _add_transaction_or_journal_data(self):
        df_data = (
            self._df_transactions
            if self.uses_transaction_report_data()
            else self._df_journals
        )
        account_col_name = "Ledger Account"
        if not self.uses_transaction_report_data():
            # The column name is slightly different in the journal data
            account_col_name += " by Identifier"

        handled_ledger_accounts = set()
        # handled_ledger_accounts.add(6400)  # F&A individual items
        handled_ledger_accounts.add(4200)  # Award revenue

        if not self.uses_transaction_report_data():
            # If we aren't using transaction data, then we are using separate journal/payroll data,
            # so remove the Salaries and Wages category from the journal data, as this will be in the
            # payroll data.
            handled_ledger_accounts.add(5000)

        # Loop through rows in the actuals DataFrame
        for index, row in self._df.iterrows():
            # Skip sub-rows (those starting with ROW_INDENT_STR)
            if (
                row[_ColumnHeaders.CATEGORY.value.col_name].startswith(ROW_INDENT_STR)
                or row[_ColumnHeaders.CATEGORY.value.col_name] == "Total"
            ):
                continue

            # Get the ledger account number from the row
            ledger_account = int(row[_ColumnHeaders.CATEGORY.value.col_name][:4])

            # Skip these
            if ledger_account in handled_ledger_accounts:
                continue
            handled_ledger_accounts.add(ledger_account)

            # Find transaction entries for this ledger account
            transactions = df_data[
                (df_data[account_col_name] == int(ledger_account))
            ].copy()

            # If using journals, filter out ADM (administration) transactions
            if not self.uses_transaction_report_data():
                transactions = transactions[
                    ~transactions["Cost Center Name"].str.startswith("[ADM]")
                ]
            if transactions.empty:
                continue

            # Clean the 'Amount' column
            transactions["Amount"] = transactions["Amount"].apply(
                workday_str_amount_to_float
            )

            # Ensure Accounting Date is datetime
            transactions["Accounting Date"] = pd.to_datetime(
                transactions["Accounting Date"], errors="coerce"
            )

            # Add a 'Month' column with abbreviation (e.g., 'Jan', 'Feb'), and 'month_idx'
            transactions["Month"] = transactions["Accounting Date"].dt.month.apply(
                lambda m: calendar.month_abbr[m]
            )
            transactions["month_idx"] = transactions["Accounting Date"].dt.month

            # Create month columns and initialize them to None or 0.0
            for month in MONTH_COLS:
                transactions[month] = None

            def assign_amount(row):
                row[row["Month"]] = row["Amount"]
                return row

            transactions = transactions.apply(assign_amount, axis=1)

            # If ledger account is 5000, then this is employee salary data, and we
            # want to group by employee and month
            if ledger_account == 5000:
                # Group by employee and month
                transactions = (
                    transactions.groupby("Employee Name 1")[list(MONTH_COLS)]
                    .sum()
                    .reset_index()
                )

                # Replace empty employee names with "(Unknown)"
                transactions["Employee Name 1"] = transactions[
                    "Employee Name 1"
                ].replace("", "(Unknown)")

                for month in MONTH_COLS:
                    transactions[month] = transactions[month].apply(
                        lambda x: "" if x == 0.0 else x
                    )
            else:
                # Sort by month_idx
                transactions = transactions.sort_values(by="month_idx")

            rows = []
            for _, transaction_row in transactions.iterrows():
                if ledger_account == 5000:
                    info_columns = ["Employee Name 1"]
                else:
                    info_columns = [
                        "Employee Name 1",
                        "Spend Category",
                        "Business Reason",
                        "Line Memo",
                        "Operational Transaction",
                    ]
                info_columns = [
                    col_name for col_name in info_columns if col_name in transaction_row
                ]

                row_data = {col_name: None for col_name in _get_column_headers()}
                row_data[_ColumnHeaders.CATEGORY.value.col_name] = (
                    ROW_INDENT_STR
                    + ", ".join(
                        filter(
                            lambda x: x and x.strip(),
                            [transaction_row[c] for c in info_columns],
                        )
                    ).strip()
                )
                for month in MONTH_COLS:
                    row_data[month] = transaction_row[month]
                # if month in row_data:
                #     row_data[month] = transaction_row["Amount"]

                rows.append(row_data)

            transaction_df = pd.DataFrame(rows, columns=_get_column_headers())

            # Add the transaction entries DataFrame as sub-rows to the actuals DataFrame
            if not transaction_df.empty:
                self._insert_subrows(
                    transaction_df, row[_ColumnHeaders.CATEGORY.value.col_name]
                )

        # Get all unhandled ledger accounts
        unhandled_ledger_accounts = (
            set(df_data[account_col_name].unique()) - handled_ledger_accounts
        )

        if unhandled_ledger_accounts:
            warn(
                "Your transaction data contains ledger accounts that are not correctly matched to the budget categories: "
                f"{', '.join(map(str, unhandled_ledger_accounts))}. "
                f"Please reach out to {DEVELOPER_NAME} to add support for these accounts."
            )

    def _add_payroll_data(self):
        """Add payroll data to the account DataFrame."""

        payroll_df = self._df_payroll

        ##################### Process payroll_df #####################
        print(f"There are {len(payroll_df)} payroll entries.")

        # Filter down to relevant grant, where 'Grant' starts with self.account_code
        if self.account_code.startswith("GR"):
            payroll_df = payroll_df[
                payroll_df["Grant"].str.startswith(f"{self.account_code} ")
            ].copy()
        elif self.account_code.startswith("AC"):
            payroll_df = payroll_df[
                payroll_df["Activity"].str.startswith(f"{self.account_code} ")
            ].copy()
        elif self.account_code.startswith("GF"):
            payroll_df = payroll_df[
                payroll_df["Gift"].str.startswith(f"{self.account_code} ")
            ].copy()
        else:
            raise ValueError("Unsupported account code for payroll data.")
        print(f"There are {len(payroll_df)} payroll entries for {self.account_code}.")

        # Add a 'Month' column based on 'Budget Date'
        payroll_df["Month"] = pd.to_datetime(payroll_df["Budget Date"]).dt.month

        # Clean the 'Amount' column
        payroll_df["Amount"] = payroll_df["Amount"].apply(workday_str_amount_to_float)

        # Step 2: Group by Employee and Month
        payroll_summary = (
            payroll_df.groupby(["Employee", "Month"])["Amount"].sum().reset_index()
        )

        # Step 3: Pivot: one row per Employee, one column per month
        # and format the month names
        payroll_pivot = payroll_summary.pivot(
            index="Employee", columns="Month", values="Amount"
        ).fillna(0)
        month_cols = [datetime(1900, m, 1).strftime("%b") for m in range(1, 13)]
        payroll_pivot.columns = [
            datetime(1900, int(m), 1).strftime("%b") for m in payroll_pivot.columns
        ]

        # Convert the data to float
        payroll_pivot = payroll_pivot.astype(float)

        # Make sure all months are there
        for m in month_cols:
            if m not in payroll_pivot.columns:
                payroll_pivot[m] = 0.0

        # Prepare the df for merging with actuals_df
        payroll_pivot[_ColumnHeaders.BUDGET.value.col_name] = ""
        payroll_pivot[_ColumnHeaders.PREV_ACTUALS.value.col_name] = ""
        payroll_pivot[_ColumnHeaders.BALANCE.value.col_name] = ""
        payroll_pivot[_ColumnHeaders.COMMITTED.value.col_name] = ""

        ytd = payroll_pivot[month_cols].sum(axis=1).astype(float)
        month_vals = payroll_pivot[month_cols].astype(float)
        payroll_pivot[month_cols] = month_vals.where(month_vals != 0.0, "")
        payroll_pivot[_ColumnHeaders.ACTUALS_YTD.value.col_name] = ytd

        # Intent employee names a bit
        payroll_pivot[_ColumnHeaders.CATEGORY.value.col_name] = payroll_pivot.index.map(
            lambda x: f"{ROW_INDENT_STR}{x}"
        )
        payroll_pivot = payroll_pivot[_get_column_headers()]
        print(
            f"There is data for {len(payroll_pivot)} employees within these payroll entries."
        )
        if payroll_pivot.empty:
            warn(
                "No payroll data found for this account. "
                "Skipping payroll data addition."
            )
            return

        # Merge with the actuals_df
        self._insert_subrows(payroll_pivot, "5000:Salaries and Wages")

    def _insert_subrows(self, df_subrows, after_ledger_account):
        # Drop columns that are all-NA
        df_subrows = df_subrows.dropna(axis=1, how="all")

        category_indices = self._df.index[
            self._df[_ColumnHeaders.CATEGORY.value.col_name] == after_ledger_account
        ]
        if category_indices.empty:
            error(
                f"Could not find the category ('{after_ledger_account}') in the DataFrame."
            )
        category_index = category_indices[0]

        before = self._df.iloc[: category_index + 1]
        after = self._df.iloc[category_index + 1 :]
        self._df = pd.concat([before, df_subrows, after], ignore_index=True)

    ############################################################################
    ##################### Methods to export to Excel and format ################
    ############################################################################

    def to_excel(self, output_path=None):
        if output_path is None:
            output_path = f"{self.account_code}_{self.year}.xlsx"
        print_color(TermColors.GREEN, f"Creating {output_path}")

        formatted_df = self._df.copy()

        formatted_df.to_excel(output_path, index=False, engine="openpyxl")
        self._format_excel(output_path)

    def _format_excel(self, excel_filename):
        """Format the Excel file with specific styles."""
        wb = load_workbook(excel_filename)
        ws = wb.active
        currency_format = "#,##0.00"

        # Insert Title Row
        ws.insert_rows(1)
        title_text = f"{self.account_code} - Year {self.year} (Generated {datetime.now().strftime('%b %d, %Y')})"
        ws.cell(row=1, column=1).value = title_text

        num_columns = ws.max_column
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Define a thin black border
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # === Insert new header row ===
        ws.insert_rows(2)  # Insert a new row 2
        for col, header_text in enumerate(_get_column_headers(), start=1):
            ws.cell(row=3, column=col).value = header_text

        # Add "Actuals" merged across Jan to YTD
        first_actuals_col = _ColumnHeaders.JAN.value.index  # 4
        last_actuals_col = _ColumnHeaders.ACTUALS_YTD.value.index  # 16
        ws.merge_cells(
            start_row=2,
            start_column=first_actuals_col,
            end_row=2,
            end_column=last_actuals_col,
        )
        actuals_cell = ws.cell(row=2, column=first_actuals_col)
        actuals_cell.value = _ColumnHeaders.ACTUALS_YTD.value.col_name
        actuals_cell.alignment = Alignment(horizontal="center", vertical="center")
        actuals_cell.font = Font(bold=True)

        # Apply borders to the merged "Actuals" cells
        for col in range(first_actuals_col, last_actuals_col + 1):
            ws.cell(row=2, column=col).border = thin_border

        # Merge vertically: Category, Budget, Prev Actuals, Balance
        for col in [
            _ColumnHeaders.CATEGORY,
            _ColumnHeaders.BUDGET,
            _ColumnHeaders.PREV_ACTUALS,
            _ColumnHeaders.COMMITTED,
            _ColumnHeaders.BALANCE,
        ]:
            ws.merge_cells(
                start_row=2,
                start_column=col.value.index,
                end_row=3,
                end_column=col.value.index,
            )
            merged_cell = ws.cell(row=2, column=col.value.index)
            merged_cell.value = col.value.col_name
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            merged_cell.font = Font(bold=True)

            # Apply borders to the merged cells
            ws.cell(row=2, column=col.value.index).border = thin_border
            ws.cell(row=3, column=col.value.index).border = (
                thin_border  # Also border the lower merged part
            )

        # Apply borders to all header cells in row 3
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=3, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Set all header cells to have wrap text
        for col in range(1, num_columns + 1):
            header_cell = ws.cell(row=2, column=col)
            header_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Style the second header row (row 3)
        for col in range(1, num_columns + 1):
            header_cell = ws.cell(row=3, column=col)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.font = Font(bold=True)

        # Freeze top 3 rows
        ws.freeze_panes = ws["B4"]

        # === Format the data rows ===
        fill_5 = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        fill_15 = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        fill_total = PatternFill(
            start_color="595959", end_color="595959", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        rows_to_group = []
        shade_toggle = True

        for idx, row in enumerate(
            ws.iter_rows(min_row=4, min_col=1), start=4
        ):  # Data now starts at row 4
            category_cell = row[0]
            is_employee = isinstance(
                category_cell.value, str
            ) and category_cell.value.startswith(ROW_INDENT_STR)

            if not is_employee:
                fill = fill_5 if shade_toggle else fill_15
                for cell in row:
                    cell.fill = fill
                shade_toggle = not shade_toggle
            else:
                rows_to_group.append(idx)

            for cell in row[1:]:
                cell.number_format = currency_format

        if rows_to_group:
            # Group the rows with the row above
            ws.sheet_properties.outlinePr.summaryBelow = False

            sorted_rows = sorted(rows_to_group)
            start = sorted_rows[0]
            end = start

            for row in sorted_rows[1:]:
                if row == end + 1:
                    # Continue the block
                    end = row
                else:
                    # End current block, start a new one
                    ws.row_dimensions.group(start, end, outline_level=1, hidden=True)
                    start = end = row

            # Group the last block
            ws.row_dimensions.group(start, end, outline_level=1, hidden=True)

        ws.sheet_format = SheetFormatProperties(outlineLevelRow=1)

        # Set column widths
        ws.column_dimensions[
            get_column_letter(_ColumnHeaders.CATEGORY.value.index)
        ].width = 50
        ws.column_dimensions[
            get_column_letter(_ColumnHeaders.BUDGET.value.index)
        ].width = 13
        ws.column_dimensions[
            get_column_letter(_ColumnHeaders.PREV_ACTUALS.value.index)
        ].width = 13
        ws.column_dimensions[
            get_column_letter(_ColumnHeaders.ACTUALS_YTD.value.index)
        ].width = 13
        ws.column_dimensions[
            get_column_letter(_ColumnHeaders.COMMITTED.value.index)
        ].width = 13
        ws.column_dimensions[
            get_column_letter(_ColumnHeaders.BALANCE.value.index)
        ].width = 13
        for month in range(
            _ColumnHeaders.JAN.value.index, _ColumnHeaders.DEC.value.index + 1
        ):
            ws.column_dimensions[get_column_letter(month)].width = 12

        # Highlight the total row
        total_row = len(ws["A"])
        for cell in ws[total_row]:
            cell.fill = fill_total
            cell.font = Font(bold=True, color="FFFFFF")

        balance_cell = ws.cell(row=total_row, column=_ColumnHeaders.BALANCE.value.index)
        balance_cell.fill = yellow_fill
        balance_cell.font = Font(bold=True, color="000000")

        wb.save(excel_filename)
