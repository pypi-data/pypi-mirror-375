import copy
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from .yutils import TermColors, print_color


def xlsx_merge(input_files, output_file):
    """Merge multiple Excel files into one, preserving styles and merged cells."""
    print_color(
        TermColors.GREEN,
        f"Merging {[str(s) for s in input_files]} files into {output_file}...",
    )
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "Merged"
    current_row_idx = 1

    output_ws.sheet_properties.outlinePr.summaryBelow = False
    output_ws.sheet_properties.outlinePr.applyStyles = True

    for idx, file_path in enumerate(input_files):
        wb = load_workbook(file_path)
        ws = wb.active

        if idx == 0:
            for col_letter, dim in ws.column_dimensions.items():
                output_ws.column_dimensions[col_letter].width = dim.width

        row_offset = current_row_idx - 1

        # Track merged ranges for post-formatting
        merged_ranges = []

        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            new_range = (
                f"{get_column_letter(min_col)}{min_row + row_offset}:"
                f"{get_column_letter(max_col)}{max_row + row_offset}"
            )
            output_ws.merge_cells(new_range)
            merged_ranges.append(((min_col, min_row, max_col, max_row), row_offset))

        for src_row in ws.iter_rows():
            src_row_idx = src_row[0].row
            dst_row_idx = current_row_idx

            for cell in src_row:
                if isinstance(cell, MergedCell):
                    continue

                dst_cell = output_ws.cell(
                    row=dst_row_idx, column=cell.column, value=cell.value
                )

                if cell.has_style:
                    dst_cell.font = copy.copy(cell.font)
                    dst_cell.border = copy.copy(cell.border)
                    dst_cell.fill = copy.copy(cell.fill)
                    dst_cell.number_format = copy.copy(cell.number_format)
                    dst_cell.protection = copy.copy(cell.protection)
                    dst_cell.alignment = copy.copy(cell.alignment)

            # Grouping & collapsed state
            src_dim = ws.row_dimensions[src_row_idx]
            dst_dim = output_ws.row_dimensions[dst_row_idx]
            dst_dim.outlineLevel = src_dim.outlineLevel
            dst_dim.hidden = src_dim.hidden

            current_row_idx += 1

        # Reapply styles to all cells in merged regions
        for (min_col, min_row, max_col, max_row), offset in merged_ranges:
            src_cell = ws.cell(row=min_row, column=min_col)
            for row in range(min_row + offset, max_row + offset + 1):
                for col in range(min_col, max_col + 1):
                    cell = output_ws.cell(row=row, column=col)
                    if src_cell.has_style:
                        cell.font = copy.copy(src_cell.font)
                        cell.border = copy.copy(src_cell.border)
                        cell.fill = copy.copy(src_cell.fill)
                        cell.number_format = copy.copy(src_cell.number_format)
                        cell.protection = copy.copy(src_cell.protection)
                        cell.alignment = copy.copy(src_cell.alignment)

        if idx < len(input_files) - 1:
            output_ws.row_dimensions[current_row_idx].height = 40  # Tall gap row
            current_row_idx += 1

    output_wb.save(output_file)
