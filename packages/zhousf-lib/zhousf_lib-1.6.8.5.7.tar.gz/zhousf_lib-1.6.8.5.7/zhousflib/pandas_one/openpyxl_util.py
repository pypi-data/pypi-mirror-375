# -*- coding: utf-8 -*-
# @Author  : zhousf
# @Function:
from pathlib import Path

import openpyxl

from zhousflib.util import re_util


def unmerge_and_fill_cells(excel_file: Path, delete_duplicates_rate: float = 1.0, tmp_excel: Path = None,
                           target_sheet_name=None) -> Path:
    """
    拆分合并单元格并填充有效值
    :param excel_file:
    :param delete_duplicates_rate: 对拆分合并单元格的结果去重的比例，默认为1.0（全相同时去重）,0则不去重
    :param tmp_excel: 临时文件，若为空则会更新源文件
    :param target_sheet_name: None第一张表
    :return:
    """
    wb = openpyxl.load_workbook(str(excel_file))
    contain_merge_cells = False
    for sheet_name in wb.sheetnames:
        if target_sheet_name:
            if target_sheet_name != sheet_name:
                continue
        worksheet = wb[sheet_name]
        all_merged_cell_ranges = list(worksheet.merged_cells.ranges)
        rows_deal = {}
        """
        拆分合并单元格
        """
        for merged_cell_range in all_merged_cell_ranges:
            merged_cell = merged_cell_range.start_cell
            worksheet.unmerge_cells(range_string=merged_cell_range.coord)
            start, end = merged_cell_range.coord.split(":")
            start = int(re_util.get_digit_char(start))
            end = int(re_util.get_digit_char(end))
            if (start, end) not in rows_deal:
                rows_deal[(start, end)] = 1
            else:
                rows_deal[(start, end)] += 1

            for row_index, col_index in merged_cell_range.cells:
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = merged_cell.value
        """
        找到符合拆分合并单元格条件的单元格rows
        """
        need_fill = []
        for i in rows_deal:
            need_fill.append(i)
            contain_merge_cells = True
        if len(need_fill) > 0:
            need_fill.sort(key=lambda x: x[0], reverse=False)
        """
        拆分合并单元格后，对空单元格赋予有效值，仅对两个的合并单元格
        """
        for cells in worksheet.iter_rows():
            for cell in cells:
                row = cell.row
                column = cell.column
                for fill in need_fill:
                    count = rows_deal.get(fill)
                    if not count:
                        continue
                    if row == fill[0] and abs(fill[1] - fill[0]) == 1:
                        next_cell = worksheet.cell(row=fill[1], column=column)
                        if not cell.value and next_cell.value:
                            cell.value = next_cell.value
                        rows_deal.pop(fill)
        """
        拆分合并单元格后会有重复的两条，这里去重一下
        """
        if delete_duplicates_rate > 0:
            # 偏移量，记录删除row的个数
            offset = 0
            for fill in need_fill:
                for i in range(fill[0], fill[1]+1):
                    current_data = []
                    next_data = []
                    if i - offset < 1:
                        continue
                    for row_cells in worksheet.iter_rows(min_row=i-offset, max_row=i-offset):
                        current_data = [cell.value for cell in row_cells]
                    if i < fill[1]:
                        for row_cells in worksheet.iter_rows(min_row=i+1-offset, max_row=i+1-offset):
                            next_data = [cell.value for cell in row_cells]
                    if len(next_data) > 0 and len(current_data) > 0:
                        # 若下一行完全包含上一行或占比满足条件，则删除上一行
                        can_overwrite = True
                        same_cell_count = 0
                        available_cell_count = 0
                        for cell in current_data:
                            if not cell:
                                continue
                            available_cell_count += 1
                            if cell not in next_data:
                                can_overwrite = False
                            else:
                                same_cell_count += 1
                        if can_overwrite or same_cell_count >= delete_duplicates_rate * available_cell_count:
                            worksheet.delete_rows(idx=i-offset)
                            offset += 1
    if tmp_excel:
        wb.save(str(tmp_excel))
        wb.close()
        return tmp_excel
    else:
        if contain_merge_cells:
            wb.save(str(excel_file))
        wb.close()
        return excel_file
