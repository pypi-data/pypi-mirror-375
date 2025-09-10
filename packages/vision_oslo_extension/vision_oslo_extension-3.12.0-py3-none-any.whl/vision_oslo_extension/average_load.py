#
# -*- coding: utf-8  -*-
#=================================================================
# Created by: Jieming Ye
# Created on: Feb 2024
# Last Modified: Feb 2024
#=================================================================
# Copyright (c) 2024 [Jieming Ye]
#
# This Python source code is licensed under the 
# Open Source Non-Commercial License (OSNCL) v1.0
# See LICENSE for details.
#=================================================================
"""
Pre-requisite: 
SimulationName.oof: default oslo output file after successfully run a simulation.
xxxxx.xlsx: excel spreadsheet with proper user settings.
SimulationName.lst.txt: required for option 4
Used Input:
simname: to locate the result file or file rename
time_start: to define the analysis start time
time_end: to define the analysis end time
option_select: to define the option selected in subfunction
text_input: to locate the excel file name
other input for oslo_extraction.py only.
Expected Output:
Updated excel spreadsheet with analysis result.
Description:
This script defines the process of doing average load assessment.
First, it reads the excel file and save the user input in a data frame called start_df. Then it find the related .d4 files and read information one by one and saved in a data frame list called d4dataframe. The calculation result is saved in a similar format data frame list called sumdataframe. It then doing some analysis and save the final result in an updated data frame. Final step is to save all information in the excel in a proper format. The whole process is easy to follow via reading the code.
The key manual updating part is get_result_range() function which depends on the desired output format in excel, followed by table_formatting() function where some reading needs manual input. The other process should be auto adjustable as long as the excel input format is followed.

"""
#=================================================================
# VERSION CONTROL
# V1.0 (Jieming Ye) - Initial Version
#=================================================================
# Set Information Variable
# N/A
#=================================================================

# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series


# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension import model_check
from vision_oslo_extension.shared_contents import SharedMethods


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    print("")
    print("Average Power Assessment - - - > ")
    print("")
    
    simname = simname
    # Specify Excel file name
    excel_file = 'output.xlsx'
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    # node_list = [['SPF1','TSC1','TSC2'],['SPF2','TSC3','TSC4']]
    # node_list = [['SPF1','TSC1','TSC2'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4']]
    #option = "1" # Fully Restart, require oof, and list
    #option = "2" # auto, oof only, self configuration file
    #option = "3" # manual, exisiting d4 and mxn, self configuration
    #option = "4" # manual, create d4 and mxn only.

    option = option_select # 1:Fully Restart, require oof, and list

    time_increment = 5
    start = 10

    space  = 5

    if option not in ["0","1","2","3","4"]:
        SharedMethods.print_message("ERROR: Error in average_load.py. Please contact Support.","31")
        return False
    
    if option == "0":
        SharedMethods.print_message("ERROR: Please select an option to proceed.","31")
        return False
    
    if option == "1":
        lstfilename = simname + '.lst.txt'
        if not SharedMethods.check_existing_file(lstfilename):
            return False
    
    if option in ["1","2","4"]:
        if not SharedMethods.check_oofresult_file(simname):
            return False

    # start_cell = 'B11'
    # read data from start tab
    # start_df = pd.read_excel(writer,sheet_name = "Start")
    result = start_reading_process(
        simname, time_start, time_end, text_input, low_v, high_v, time_step,excel_file,option
    )
    if result == False:
        return False
    else:
        supply_list, node_list, start_df = result

    # number of dataframes to create
    total = len(supply_list)

    # check if want to go throught the feeder check process
    if option in ["1","2","4"]:

        if not one_stop_AC_extraction(simname, time_start, time_end,supply_list):
            return False

        if option == "4":
            return True
        
    if option == "3":
        print("Checking essential d4 files and mxn files...")
        for supply in supply_list:
            filename = simname + "_" + supply + ".osop.d4"
            if not SharedMethods.check_existing_file(filename):
                SharedMethods.print_message(f"ERROR: d4 file {supply} do not exist. Select option 2 to proceed.","31")
                return False
        filename = simname + ".osop.mxn"
        if not SharedMethods.check_existing_file(filename):
            SharedMethods.print_message(f"ERROR: File {filename} do not exist. Select option 2 to proceed.","31")
            return False
    
    
    #feeder_number = len(supply_list)
    try:
        # read min max value
        mxndataframe = mxn_file_reading(simname)

        # process d4 file
        p_r_df,i_r_df,d4dataframe,sumdataframe = feeder_reading_process(simname, supply_list,time_increment)
        
        # update NPS
        p_r_df = nps_update(p_r_df, start_df, supply_list)

        # update min max value voltage
        p_r_df = minmax_update(simname,p_r_df,supply_list,mxndataframe,node_list)

        # process information to get the summary (substation real time sum)
        start_sum_df,s_r_df, fsdataframe,fssum, feeder_dict = summary_info_power(supply_list,start_df,d4dataframe)


        # get feeder station list
        feeder_station = list(feeder_dict.keys())
        fsnumber = len(feeder_station)

        # update NPS
        s_r_df = nps_update(s_r_df, start_sum_df, feeder_station)

        # items_total = start_df.shape[0]
        # print(str(items_total))

        data_write_save(simname, excel_file,start,total,space,fsnumber,start_df, start_sum_df,p_r_df,i_r_df,s_r_df, \
                        mxndataframe,d4dataframe,fsdataframe,sumdataframe,fssum, \
                            supply_list,node_list,feeder_station,feeder_dict)

    # # Calculate the Excel range for each DataFrame
    # range_list = get_result_range(start,total,space,fsnumber) # range list as [power data range, current data range]
    
    
    # # table formatting
    # table_formatting(simname,excel_file,range_list,mxndataframe)
    
    except KeyError as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}. Possibly due to incompleted data.","31")
        return False
    
    except Exception as e:
        SharedMethods.print_message(f"ERROR: Unexpected error occured: {e}","31")
        return False

    return True

# extraction
def one_stop_AC_extraction(simname, time_start, time_end,supply_list):

    # User defined time windows extraction
    time_start = SharedMethods.time_input_process(time_start,1)
    time_end = SharedMethods.time_input_process(time_end,1)

    if time_start == False or time_end == False:
        return False

    # process 1: feeder step output
    print("Extrating feeder step output from Start Tab ... ")

    # processing the list
    for items in supply_list:
        #print(items)
        if not oslo_extraction.feeder_step_one(simname,items,time_start,time_end):
            SharedMethods.print_message(f"WARNING: Error for {items} will be ignored and process continued...","33")

    # process 2: minmax value extraction
    print("Extrating min-max output...")
    
    # define the default osop command
        
    opcname = simname + ".opc"

    with open(opcname,"w") as fopc:
        fopc.writelines("MINMAX VALUES REQUIRED\n")

    if not SharedMethods.osop_running(simname):
        return False

    return True

# write data to excel
def data_write_save(simname,excel_file,start,total,space,fsnumber,start_df,start_sum_df,p_r_df,i_r_df,s_r_df, \
                    mxndataframe,d4dataframe,fsdataframe,sumdataframe,fssum, \
                        supply_list,node_list,feeder_station,feeder_dict):
    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Result Page...")
        start_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)
        p_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = start_df.shape[1] + 1)
        #p_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = start_df.shape[1] + 1, float_format="%.2f")

        start_df.to_excel(writer, sheet_name="Result", index=False, startrow = start + total + space, startcol = 1)
        i_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start + total + space, startcol = start_df.shape[1] + 1)

        start_sum_df.to_excel(writer, sheet_name="Result", index=False, startrow = start + total*2 + space*2, startcol = 1)
        s_r_df.to_excel(writer, sheet_name="Result", index=False, startrow = start + total*2 + space*2, startcol = start_sum_df.shape[1] + 1)
        
        
        print("Writing min max summary...")
        mxndataframe.to_excel(writer, sheet_name='Data', index=False, startrow = len(feeder_dict)+2+len(supply_list)+2+2)
        
        # Write each DataFrame to a different sheet in the Excel file
        print("Writing individual feeder...")
        for index, dflist in enumerate(d4dataframe):
            print(f"Saving {supply_list[index]}...")
            # emty columns
            dflist.insert(dflist.columns.get_loc('I_angle')+1,'New_C_1', np.nan)
            dflist.insert(dflist.columns.get_loc('S_30min')+1,'New_C_2', np.nan)
            dflist.insert(dflist.columns.get_loc('I_1800s_RMS')+1,'New_C_3', np.nan)
            dflist.insert(dflist.columns.get_loc('PF_30min')+1,'Excel_Time', np.nan)

            sumdataframe[index].insert(sumdataframe[index].columns.get_loc('I_angle')+1,'New_C_1', np.nan)
            sumdataframe[index].insert(sumdataframe[index].columns.get_loc('S_30min')+1,'New_C_2', np.nan)
            sumdataframe[index].insert(sumdataframe[index].columns.get_loc('I_1800s_RMS')+1,'New_C_3', np.nan)

            sumdataframe[index].to_excel(writer, sheet_name=supply_list[index], index=False, startrow = 0)

            dflist.to_excel(writer, sheet_name=supply_list[index], index=False, startrow = sumdataframe[index].shape[0]+2)

            sheet = wb[supply_list[index]]
            power_plot(dflist,sheet,row = sumdataframe[index].shape[0]+2+2,option=1)

        print("Writing feeder station...")
        for index, dflist in enumerate(fsdataframe):
            fssum[index].to_excel(writer, sheet_name=feeder_station[index], index=False, startrow = 0)

            dflist.insert(dflist.columns.get_loc('30minS')+1,'Excel_Time', np.nan)

            dflist.to_excel(writer, sheet_name=feeder_station[index], index=False, startrow = fssum[index].shape[0]+2)
            
            sheet = wb[feeder_station[index]]
            power_plot(dflist,sheet,row = fssum[index].shape[0]+2+2,option=2)
        
        # update Data tab
        data_tab_update(wb,feeder_dict,supply_list,node_list)
        
        # # Calculate the Excel range for each DataFrame
        range_list = get_result_range(start,total,space,fsnumber) # range list as [power data range, current data range]
        
        # # table formatting
        table_formatting(simname,wb,range_list,mxndataframe)

        print("Saving Data...")

    return

# plot the 30min average power:
def power_plot(dflist,sheet,row,option):
    
    # # Create a new worksheet to plot the data
    # worksheet = writer.sheets[sheetname]
    if option == 1: # single supply points
        index = 8
        range = f"AX{row}:AX{row+len(dflist)-1}"
        x = 50 #  column index for time
        y = 26 #  column index for value
        for line in sheet[range]:
            for cell in line:
                cell.value = f'=TIME(LEFT(C{index},2),MID(C{index},4,2),RIGHT(C{index},2))'
                cell.number_format = 'hh:mm:ss'
            index = index + 1

    if option == 2: # feeder station tab
        index = 8
        range = f"T{row}:T{row+len(dflist)-1}"
        x = 20 #  column index for time
        y = 19 #  column index for value
        for line in sheet[range]:
            for cell in line:
                cell.value = f'=TIME(LEFT(A{index},2),MID(A{index},4,2),RIGHT(A{index},2))'
                cell.number_format = 'hh:mm:ss'
            index = index + 1


    # Create a scatter plot
    # chart = writer.book.add_chart({'type': 'scatter'})
    chart = ScatterChart('smoothMarker')
    chart.title = "30-min Average Apparent Power"
    chart.x_axis.title = "Time (hh:mm:ss)"
    chart.y_axis.title = "Average Power (MVA)"

    #time_format = NamedStyle(name='time_format', number_format='HH:MM:SS')
    # Set chart size
    chart.height = 10  # Adjust width as needed
    chart.width = 25  # Adjust height as needed

    # # Specify the range (C8 to C100) to which the style should be applied
    # for line in sheet.iter_rows(min_row=row, max_row=row+len(dflist)-1, min_col=3, max_col=3):
    #     for cell in line:
    #         cell.number_format = 'hh:mm:ss'

    xvalues = Reference(sheet, min_col=x, min_row=row, max_row=row+len(dflist)-1)


    yvalues = Reference(sheet, min_col=y, min_row=row, max_row=row+len(dflist)-1)
    series = Series(yvalues,xvalues,title_from_data=False,title = '30min Average')
    chart.series.append(series)
    chart.legend.position = 't' # top legend position

    if option == 1:
        sheet.add_chart(chart,"J8")
    if option == 2:
        sheet.add_chart(chart,"T8")
    
    # chart.add_series({
    #     'categories': f'={sheetname}!$C${row}:$C${row+len(dflist)+1}',  # Assuming 'Time' is in column B
    #     'values': f'={sheetname}!$Z${row}:$Z${row+len(dflist)+1}',  # Assuming '30-minP' is in column C
    #     'line': {'width': 1.5},  # Adjust line width as needed
    # })

    # # Set chart title and axis labels
    # chart.set_title({'name': 'Supply Point 30-min Average Apparent Power'})
    # chart.set_x_axis({'name': 'Time (hh:mm:ss)'})
    # chart.set_y_axis({'name': 'Average Power (MVA)'})

    # # Insert the chart into the worksheet
    # worksheet.insert_chart('J8', chart)  # Adjust the position as needed

    return

# read the start tab and collect informaiton
def start_reading_process(simname, time_start, time_end, text_input, low_v, high_v, time_step, excel_file,option):
    supply_list = []
    node_list = []
    name_list = []
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                table_list = check_table_list(wb[sheet_name])
                
                if table_list is not False:
                    if option != "1":
                        # get supply name list:
                        for items in table_list:
                            name_list.append(items[0])
                        node_list = check_node_list(name_list,wb[sheet_name])
                        if node_list == False:
                            return False
                    else:
                        filename = simname + ".lst.txt"
                        node_check = model_check.main_menu(simname, filename, "2", \
                                           time_start, time_end, "sp", text_input, low_v, high_v, time_step)
                        # Extract elements from the third element to the end for each inner list

                        for line in table_list:
                            supply = line[1] # get the supply point oslo id
                            for lst in node_check:
                                if lst[1] == supply:
                                    nodeline = lst[2:]
                                    break                                    
                                else:
                                    nodeline = []

                            try:
                                node_list.append(nodeline)
                                if nodeline[1] == 'NOT CONNECTED':
                                    SharedMethods.print_message(f"ERROR: {nodeline[0]} is not connected to any nodes. Double check or choose option 3 to Proceed.","31")
                                    return False

                            except Exception as e:
                                SharedMethods.print_message(f"ERROR: {e}. [Potential reason is OSLO mode is not enabled when running the simulation. Double check...]","31")
                                SharedMethods.print_message(f"ERROR: Exception happens at processing {supply}. Ensure the Supply Point OSLO ID is correct...","31")
                                return False

                        #node_list = [lst[2:] for lst in node_check]
                        #node_list = [['SPF1','TSC1','TSC2'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4'],['SPF2','TSC3','TSC4']]
                else:
                    return False
            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)

    except Exception as e:
        SharedMethods.print_message(f"(Close the Excel file and Start Again) Error: {e}","31")
        return False
        
    # read start up information 
    columns = ["Supply Point Name","OSLO ID","Incoming Voltage (kV)","Incoming Feeder","Capacity (MVA)","Fault Level (MVA)","Feeder Station Name","Cable Rating (MVA)","FSC (MVA)","Supply Type","Supply Voltage (kV)"]
    start_df = pd.DataFrame(table_list,columns=columns)
    # get supply list oslo id
    for items in table_list:
        supply_list.append(items[1])
    
    # check duplication in OSLO id
    if not SharedMethods.find_duplicates(supply_list): return False

    return supply_list, node_list, start_df

# check table list on start page
def check_table_list(sheet):
    print("Reading Configuration Setting ...")
    table_row = 11
    table_start_column = 1
    table_end_column = 11
    # create node and feeder list
    table_list = []

    # check feeder
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column+1).value is not None:
        while True:
            row_data = []
            for temp in range(table_start_column,table_end_column + 1):
                row_data.append(sheet.cell(row=index, column=temp).value)
            table_list.append(row_data)
            index += 1
            
            check = sheet.cell(row=index, column=column+1).value
            if check is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at B11.","31")
        return False

    return table_list

# check node list
def check_node_list(supply_list,sheet):

    NODE_START_INDEX = 11
    NODE_COLUMN = 16

    node_list = []
    # create nested node_list
    for feeder in supply_list:
        node_list.append([])

    # populate node information
    index = NODE_START_INDEX
    column = NODE_COLUMN
    feederid = None

    if sheet.cell(row=index, column=column).value is not None:
        while True:
            node = sheet.cell(row=index, column=column).value
            feeder = sheet.cell(row=index, column=column - 2).value
            if feeder is not None:
                if feeder in supply_list:
                    feederid = supply_list.index(feeder)
                else:
                    SharedMethods.print_message(f"Warning: Supply Point '{feeder}' cannot be found in the table. Check Name in Column N.","33")

            if node is not None and feederid is not None:
                node_list[feederid].append(node)
            else:
                feederid = None

            index += 1
            check = sheet.cell(row=index, column=column).value
            check1 = sheet.cell(row=index + 1, column=column).value
            if check is None and check1 is None:
                break
    else:
        SharedMethods.print_message("ERROR: Wrong data format. No information at P11","31")
        return False
    
    # Check if there is any empty list in node_list
    final = False
    for i,lst in enumerate(node_list):
        if lst == []:
            SharedMethods.print_message(f"ERROR: Feeder {supply_list[i]} has no voltage node associdated. Check the input before proceed.","31")
            final = True
    
    if final == True:
        SharedMethods.print_message("ERROR:Cannot proceed with this option on voltage. Check input value.","31")
        return False

    return node_list

# update data tab with connection information
def data_tab_update(wb,feeder_dict,supply_list,node_list):
    print("Writing Connection Info...")
    sheet = wb['Data']
    row = 1

    # Wrting feeder station info
    sheet.cell(row=row, column=1).value = 'Feeder Station'
    sheet.cell(row=row, column=2).value = 'Linked Supply Points'

    for key, lst in feeder_dict.items():
        row = row + 1
        sheet.cell(row = row, column = 1).value = key
        for index, supply in enumerate(lst):
            sheet.cell(row = row, column = index + 2).value = supply
    

    # for station, supply in
    row = row + 2

    row = row + 1
    sheet.cell(row=row, column=1).value = 'Supply Points'
    sheet.cell(row=row, column=2).value = 'Linked Nodes'

    for index, supply in enumerate(supply_list):
        row = row + 1
        sheet.cell(row = row, column = 1).value = supply
        
        for index1, node in enumerate(node_list[index]):
            sheet.cell(row = row, column = index1 + 2).value = node

    return

# ceate table 1 frame / power
def power_result(supply_list):
    columns = ['Instantanous','01 Minute','02 Minutes','10 Minutes', '20 Minutes', '30 Minutes']
    rows = []

    for feeder in supply_list:
        rows.append(feeder)
    df = pd.DataFrame(index=rows,columns=columns)

    return df

# create table 2 frame/ current
def current_result(supply_list):
    columns = ['Instantanous','05 seconds','10 seconds','15 seconds', '20 seconds', '25 seconds','30 seconds','35 seconds','40 seconds', \
               '60 seconds', '120 seconds', '10 min','20 min','30 min']
    rows = []

    for feeder in supply_list:
        rows.append(feeder)
    df = pd.DataFrame(index=rows,columns=columns)

    return df

# update power based on output of d4 process
def power_update(target,df_sum,feeder):
    index = 9
    for column in target.columns:
        target.at[feeder,column] = df_sum.iloc[0,index]
        index = index + 3
    return target

# update current based on output of d4 process
def current_update(target,df_sum,feeder):
    index = 25
    for column in target.columns:
        target.at[feeder,column] = df_sum.iloc[0,index]
        index = index + 1
    return target

# update power data frame inlcude the NPS summary value
def nps_update(target,start_df,supply_list):

    print("Calculating NPS....")
    #columns = ['Instantanous','01 Minute','02 Minutes','10 Minutes', '20 Minutes', '30 Minutes']
    columns = ['01 Minute','10 Minutes','30 Minutes']
    rows = []

    for feeder in supply_list:
        rows.append(feeder)
    nps_df = pd.DataFrame(index=rows,columns=columns)

    for index, feeder in enumerate(supply_list):
        #nps_df['Instantanous'].at[feeder] = target['Instantanous'].at[feeder] / start_df.iloc[index,5]
        # Check if '01 Minute' value is numeric
        if pd.notna(pd.to_numeric(target['01 Minute'].at[feeder], errors='coerce')):
            nps_df['01 Minute'].at[feeder] = target['01 Minute'].at[feeder] / start_df.iloc[index,5]
        #nps_df['02 Minutes'].at[feeder] = target['02 Minutes'].at[feeder]/ start_df.iloc[index,5]
        if pd.notna(pd.to_numeric(target['10 Minutes'].at[feeder], errors='coerce')):
            nps_df['10 Minutes'].at[feeder] = target['10 Minutes'].at[feeder]/ start_df.iloc[index,5]
        #nps_df['20 Minutes'].at[feeder] = target['20 Minutes'].at[feeder]/ start_df.iloc[index,5]
        if pd.notna(pd.to_numeric(target['30 Minutes'].at[feeder], errors='coerce')):
            nps_df['30 Minutes'].at[feeder] = target['30 Minutes'].at[feeder]/ start_df.iloc[index,5]
    
    combined_df = pd.concat([target, nps_df], axis=1)

    return combined_df

# update min max value
def minmax_update(simname,target,supply_list,mxndataframe,node_list):

    print("Caculating Minimum Maximum Value...")
    # Create a DataFrame to store the results
    
    columns=['Feeder', "Node", "Minimum", "Time at Min", "Maximum", "Time at Max"]
    rows = []
    for feeder in supply_list:
        rows.append(feeder)

    result_df = pd.DataFrame(index=rows,columns=columns)

    # Find minimum voltage for each node
    for index, feeder in enumerate(supply_list):
        columns=["Node","Minimum", "Time at Min", "Maximum", "Time at Max"]
        rows = []
        for node in node_list[index]:
            rows.append(node)

        temp_df = pd.DataFrame(index=rows,columns=columns)

        for node in node_list[index]:

            # Filter mxndataframe for the current feeder and node
            node_data = mxndataframe[(mxndataframe['Node'] == node)].reset_index(drop=True)
            
            if not node_data.empty:
                # Find the row with the minimum voltage
                temp_df.at[node,'Node'] = node_data['Node'][0]
                temp_df.at[node,'Minimum'] = node_data['Minimum Voltage (kV)'][0]
                temp_df.at[node,'Time at Min'] = node_data['Time at Min'][0]
                temp_df.at[node,'Maximum'] = node_data['Maximum Voltage (kV)'][0]
                temp_df.at[node,'Time at Max'] = node_data['Time at Max'][0]
        
        # update the main dataframe
        temp_df = temp_df.reset_index(drop=True)
        min_index = temp_df['Minimum'].idxmin()       
        result_df.at[feeder,'Feeder'] = feeder
        result_df.at[feeder,'Node'] = temp_df.at[min_index,'Node']
        result_df.at[feeder,'Minimum'] = temp_df.at[min_index,'Minimum']
        result_df.at[feeder,'Time at Min'] = temp_df.at[min_index,'Time at Min']
        result_df.at[feeder,'Maximum'] = temp_df.at[min_index,'Maximum']
        result_df.at[feeder,'Time at Max'] = temp_df.at[min_index,'Time at Max']

    # Selecting the relevant columns from result_df
    selected_columns = result_df[['Minimum', 'Maximum']]

    # Concatenate the selected columns to the 'target' DataFrame
    combined_df = pd.concat([target, selected_columns], axis = 1)

    return combined_df

# read individual d4 file
def feeder_reading_process(simname, supply_list,time_increment):

    # supply_list = ["SUP1","SUP2"]
    p_r_df = power_result(supply_list)
    i_r_df = current_result(supply_list)
    
    # create dataframe
    d4dataframe = []
    sumdataframe = []

    for feeder in supply_list:
        print(f"Processing {feeder} ...")
        filename = simname + "_" + feeder +".osop.d4"
        delimiter = '\\s+'
        columns = ["SupplyID","Type","Time","P_inst","Q_inst","Voltage","V_angle","Current","I_angle"]
        
        dtype_mapping = {"Time": str,}

        df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping)  # You might need to adjust the delimiter based on your file
        #df["Time"] = pd.to_datetime(df["Time"],format='%H:%M:%S')
        # Extracting parts from the string and formatting the 'Time' column
        df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")

        # data frame process
        df = d4_file_process(df,time_increment)
        # calculate sum value
        df_sum = d4_find_max(df)

        #update summary result
        p_r_df = power_update(p_r_df, df_sum, feeder)
        i_r_df = current_update(i_r_df, df_sum,feeder)

        # # Calculate rolling RMS
        # df['Rolling_RMS'] = df['Current'].rolling(window=12*time_increment).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)
        # calculate sum value
        sumdataframe.append(df_sum)
        d4dataframe.append(df)
    
    # power_data_sum = power_detail(d4dataframe,supply_list)
    # current_data_sum = current_detail(d4dataframe,supply_list)

    # columns = ['P-inst','P-1min','P-2min','P-2min','P-2min','P-2min','P-2min','P-2min','P-2min','P-2min']
    # rows = []
    return p_r_df, i_r_df, d4dataframe, sumdataframe

# doing invididual d4 file process and calculatoin
def d4_file_process(df, time_increment):

    df['S_inst'] = np.sqrt(df['P_inst']**2 + df['Q_inst']**2)

    window_sizes = {'1min': 60, '2min': 120, '10min': 600, '20min': 1200, '30min': 1800}

    for time_interval, window_size in window_sizes.items():
        df[f'P_{time_interval}'] = df['P_inst'].rolling(window=int(window_size / time_increment)).mean()
        df[f'Q_{time_interval}'] = df['Q_inst'].rolling(window=int(window_size / time_increment)).mean()
        df[f'S_{time_interval}'] = np.sqrt(df[f'P_{time_interval}']**2 + df[f'Q_{time_interval}']**2)

    df['I_Inst'] = df['Current'] # Current Duplicate

    for time_interval in [5, 10, 15, 20, 25, 30, 35, 40, 60, 120, 600, 1200, 1800]:
        column_name = f'I_{time_interval}s_RMS'
        df[column_name] = calculate_rolling_rms(df, 'Current', time_interval, time_increment)

    df['P_CosPhi'] = df['P_inst'] / df['S_inst']
    
    for time_interval in ['inst', '1min', '2min', '10min', '20min', '30min']:
        df[f'PF_{time_interval}'] = df[f'P_{time_interval}'] / df[f'S_{time_interval}']

    return df

# calculate rms
def calculate_rolling_rms(data, column, window_size, time_increment):
    return data[column].rolling(window=int(window_size / time_increment)).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)

# find maximum value and time of each d4 file
def d4_find_max(df):
    # # Insert two empty rows above the first row
    # df = pd.concat([pd.DataFrame(index=range(2)), df], ignore_index=True)
    sum_df = pd.DataFrame(columns=df.columns,index=range(4))
    sum_df.iloc[0, 0] = "Maximum Value (absolute for PF)"
    sum_df.iloc[1, 0] = "Maximum Value at Time"
    sum_df.iloc[2, 0] = "Minimum Value (absolute for PF)"
    sum_df.iloc[3, 0] = "Minimum Value at Time"


    if df.empty:
        sum_df.iloc[0, 1] = "DATA FOR THIS FEEDER IS NOT AVAILABLE"

    else:
        for column in df.columns[9:25]:
            if df[column].dtype in [int, float]:
                max_value = df[column].max()
                time_of_max = df.loc[df[column].idxmax(), 'Time']
                sum_df.at[0, column] = max_value
                sum_df.at[1, column] = time_of_max
                min_value = df[column].min()
                time_of_min = df.loc[df[column].idxmin(), 'Time']
                sum_df.at[2, column] = min_value
                sum_df.at[3, column] = time_of_min
        
        for column in df.columns[25:]:
            if df[column].dtype in [int, float]:
                # Check if the column has any non-NA values
                if df[column].dropna().empty:
                    time_of_max = None  # or handle this case as needed
                else:
                    max_value = abs(df[column]).max()
                    time_of_max = df.loc[abs(df[column]).idxmax(), 'Time']
                    sum_df.at[0, column] = max_value
                    sum_df.at[1, column] = time_of_max
                    min_value = abs(df[column]).min()
                    time_of_min = df.loc[abs(df[column]).idxmin(), 'Time']
                    sum_df.at[2, column] = min_value
                    sum_df.at[3, column] = time_of_min

    return sum_df

# read the mxn file and find the node voltage
def mxn_file_reading(simname):
    # Create an empty DataFrame with the specified columns
    columns = ["Node", "Minimum Voltage (kV)", "Time at Min", "Maximum Voltage (kV)", "Time at Max"]
    mxndataframe = pd.DataFrame(columns=columns)

    print(f"Processing {simname} ...")
    filename = simname + ".osop.mxn"
    

    # Specify column widths based on your file format
    colspecs = [(0, 16), (16, 38), (38, 58), (58, 73), (73, 120)]

    start_marker = "NODE VOLTAGES"
    end_marker = "MAXIMUM FEEDER"

    # Use pd.read_fwf to read the fixed-width file
    mxndataframe = pd.read_fwf(filename, colspecs=colspecs, header=None, names=columns, skiprows=2)

    # Identify the start and end indices of the relevant block
    start_index = mxndataframe.index[mxndataframe['Node'] == start_marker].tolist()[0] + 5 #(do not include NODE + ========line)
    end_index = mxndataframe.index[mxndataframe['Node'] == end_marker].tolist()[0]

    # Extract the relevant block
    mxndataframe = mxndataframe.iloc[start_index:end_index].reset_index(drop=True)

    # Strip leading and trailing spaces from all values
    mxndataframe = mxndataframe.map(lambda x: x.strip() if isinstance(x, str) else x)

    # Convert the time columns to the desired format
    time_columns = ["Time at Min", "Time at Max"]
    mxndataframe[time_columns] = mxndataframe[time_columns].map(convert_time_format)

    # Convert "Minimum Voltage (kV)" and "Maximum Voltage (kV)" to numeric
    numeric_columns = ["Minimum Voltage (kV)", "Maximum Voltage (kV)"]
    mxndataframe[numeric_columns] = mxndataframe[numeric_columns].apply(pd.to_numeric, errors='coerce')

    # Drop rows where all values are NaN or empty
    mxndataframe = mxndataframe.dropna(how='all')

    # flag = False

    # with open(filename) as fp:
    #     for line in fp:

    #         if line[:20].strip() == "NODE":
    #             flag = True
    #             continue
    #         if line[:9].strip() == "MAXIMUM":
    #             flag = False
    #             break

    #         if flag == True:
    #             if line[:20].strip() == "":
    #                 continue
    #             else:
    #                 node = line[:20].strip()
    #                 minvalue = line[21:37].strip()
    #                 tmin = line[38:55].strip()
    #                 maxvalue = line[56:71].strip()
    #                 tmax = line[72:90].strip()

    #                 data = [node,minvalue,tmin,maxvalue,tmax]

    #                 row_data = dict(zip(columns, data))
    #                 mxndataframe = mxndataframe.append(row_data, ignore_index=True)


            # Assuming the data in each line is separated by some delimiter (e.g., comma or tab)
            # data = line.strip().split('\t')  # Adjust the delimiter as needed

                    # Assuming the data has the same format as the columns
            #if len(data) == len(columns):
                        

    #df = pd.read_csv(filename, delimiter=delimiter, names = columns, skiprows = 11,dtype = dtype_mapping)  # You might need to adjust the delimiter based on your file
    #df["Time"] = pd.to_datetime(df["Time"],format='%H:%M:%S')
    # Extracting parts from the string and formatting the 'Time' column
    #df['Time'] = df['Time'].apply(lambda x: f"{int(x[1:3]):02d}:{int(x[3:5]):02d}:{int(x[5:]):02d}")
    
    return mxndataframe

# covert mxn time format to proper time format (ignore DAY)
def convert_time_format(time_str):

    if ':' not in time_str:
        return time_str  # Return the original string if ':' is not present

    # Split the time string into components
    days, time = time_str.split(':')
    
    # Split the time component into hours, minutes, and seconds
    hours, minutes, seconds = map(int, time.split('.'))

    # Convert to the desired format
    converted_time = f'{hours:02d}:{minutes:02d}:{seconds:02d}'

    return converted_time

# Table 3 creationsupply_list
def summary_info_power(supply_list,start_df,d4dataframe):
    print("Creating summary table for feeder station...")

    # Create a new DataFrame called start_sum_df with the same columns as start_df
    start_sum_df = pd.DataFrame(columns=start_df.columns)

    # Keep track of seen feeder stations
    feeder_dict = {}

    # Check feeder in supply_list one by one
    for feeder in supply_list:
        # Filter the start_df column 'OSLO feeder name' matching the feeder
        feeder_data = start_df[start_df['OSLO ID'] == feeder]

        # Check if 'Feeder Station' name in start_df has never appeared before
        feeder_station_name = feeder_data['Feeder Station Name'].iloc[0]
        if feeder_station_name not in feeder_dict:
            # If the feeder station is not in the dictionary, add it with the OSLO ID as a list
            feeder_dict[feeder_station_name] = [feeder]
            # If start_sum_df is empty, assign the columns from feeder_data
            if start_sum_df.empty:
                start_sum_df = feeder_data.copy()
            else:
                # Add the entire row information to start_sum_df
                start_sum_df = pd.concat([start_sum_df, feeder_data])
        else:
            # If the feeder station is already in the dictionary, append the OSLO ID to its list
            feeder_dict[feeder_station_name].append(feeder)
            # Update the set of seen feeder stations
            # seen_feeder_stations.add(feeder_data['Feeder Station Name'].iloc[0])
    
    # create summary result
    columns = ['Instantanous','01 Minute','02 Minutes','10 Minutes', '20 Minutes', '30 Minutes']
    row = list(feeder_dict.keys())
    # create summary result data frame
    s_r_df = pd.DataFrame(index = row, columns=columns)

    s_r_df,fsdataframe,fssum = feeder_station_sum(feeder_dict,d4dataframe,supply_list,s_r_df)

    return start_sum_df,s_r_df,fsdataframe,fssum,feeder_dict

# individual feeder staiton summary process
def feeder_station_sum(feeder_dict, d4dataframe, supply_list,s_r_df):
    fsdataframe = []
    fssum = []

    for station, oslo_list in feeder_dict.items():
        # define data frame for summary
        columns = ['Time', 'instP', 'instQ', 'instS', '1minP', '1minQ', '1minS', '2minP', '2minQ', '2minS',
                   '10minP', '10minQ', '10minS', '20minP', '20minQ', '20minS', '30minP', '30minQ', '30minS']
        df = pd.DataFrame(columns=columns)

        df['Time'] = d4dataframe[0]['Time']

        for time_interval in ['inst', '1min', '2min', '10min', '20min', '30min']:
            tempP_sum = sum(d4dataframe[supply_list.index(feeder)]['P_' + time_interval] for feeder in oslo_list)
            tempQ_sum = sum(d4dataframe[supply_list.index(feeder)]['Q_' + time_interval] for feeder in oslo_list)
            tempS_sum = np.sqrt(tempP_sum**2 + tempQ_sum**2)

            df.loc[:, f'{time_interval}P'] = tempP_sum
            df.loc[:, f'{time_interval}Q'] = tempQ_sum
            df.loc[:, f'{time_interval}S'] = tempS_sum

        # calculate sum value
        df_sum = fs_find_max(df)
        
        # update result
        index = 3
        for column in s_r_df.columns:
            s_r_df.at[station,column] = df_sum.iloc[0,index]
            index = index + 3

        fsdataframe.append(df)
        fssum.append(df_sum)

    return s_r_df,fsdataframe,fssum

# find maximum value and time of each feeder station
def fs_find_max(df):
    # # Insert two empty rows above the first row
    # df = pd.concat([pd.DataFrame(index=range(2)), df], ignore_index=True)
    sum_df = pd.DataFrame(columns=df.columns,index=range(4))
    sum_df.iloc[0, 0] = "Maximum Value (absolute for PF)"
    sum_df.iloc[1, 0] = "Maximum Value at Time"
    sum_df.iloc[2, 0] = "Minimum Value (absolute for PF)"
    sum_df.iloc[3, 0] = "Minimum Value at Time"

    for column in df.columns[1:]:
        df[column] = pd.to_numeric(df[column], errors='coerce') # convert the column to number
        if np.issubdtype(df[column].dtype, np.number):
            # Check if the entire column is NaN
            if df[column].isna().all():
                sum_df.at[0, column] = "N/A"
            else:
                max_value = df[column].max()
                time_of_max = df.loc[df[column].idxmax(), 'Time']
                sum_df.at[0, column] = max_value
                sum_df.at[1, column] = time_of_max
                min_value = df[column].min()
                time_of_min = df.loc[df[column].idxmin(), 'Time']
                sum_df.at[2, column] = min_value
                sum_df.at[3, column] = time_of_min

    # for column in df.columns[25:]:
    #     if df[column].dtype in [int, float]:
    #         max_value = abs(df[column]).max()
    #         time_of_max = df.loc[abs(df[column]).idxmax(), 'Time']
    #         sum_df.at[0, column] = max_value
    #         sum_df.at[1, column] = time_of_max
    #         min_value = abs(df[column]).min()
    #         time_of_min = df.loc[abs(df[column]).idxmin(), 'Time']
    #         sum_df.at[2, column] = min_value
    #         sum_df.at[3, column] = time_of_min
    return sum_df

def get_result_range(start,total,space,fsnumber):

    """
    Generates a list of range strings for different sections in a spreadsheet.

    Parameters:
    - start (int): Starting row index.
    - total (int): Total number of rows.
    - space (int): Space between sections.
    - fsnumber (int): Placeholder for fsnumber.

    Returns:
    List of range strings for different sections.
    """

    range_list = []
    # 0: table frame range (power, current, sum)
    # 1: table data range (power, current, sum)
    # 2: title range (2nd row)
    # 3: title range (1st row simulation info section1,2,3)
    # 4: title range (1st row power, NPS, Voltage)
    # 5: title range (1st row current)
    # 6: conditonal formatting range FSC (criteria, range)
    # 7: conditonal formatting range NPS (range, 1min, 10min, 30min)
    # 8: conditonal formatting range voltage ( min, max)
    # 9: title range (1st row power, NPS)
    # 10: conditonal formatting range FSC (criteria, range)
    # 11: conditonal formatting range NPS (range, 1min, 10min, 30min)

    
    # 0
    result_range = [f"B{start + 1}:W{start + total + 1}",f"B{start + total + space + 1}:Z{start + total*2 + space + 1}", \
               f"B{start + total*2 + space*2 + 1}:U{start + total*2 + space*2 + fsnumber + 1}"]
    range_list.append(result_range)
    # 1
    data_range = [f"M{start + 2}:W{start + total + 1}",f"M{start + total + space + 2}:Z{start + total*2 + space + 1}", \
                  f"B{start + total*2 + space*2 + 2}:U{start + total*2 + space*2 + fsnumber + 1}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"B{start+1}:W{start+1}",f"B{start + total + space + 1}:Z{start + total + space + 1}", \
                   f"B{start + total*2 + space*2 + 1}:U{start + total*2 + space*2 + 1}"]
    range_list.append(title_range)
    
    #3 4 5
    add_range_1 = [f"B{start}:L{start}",f"B{start + total + space}:L{start + total + space}",\
                   f"B{start + total*2 + space*2}:L{start + total*2 + space*2}"]
    range_list.append(add_range_1)

    add_range_2 = [f"M{start}:R{start}",f"S{start}:U{start}",f"V{start}:W{start}"]
    range_list.append(add_range_2)
    add_range_3 = [f"M{start + total + space}:Z{start + total + space}"]
    range_list.append(add_range_3)

    #6 7 8
    condi_FSC = [f"$J{start + 2}",f"$R${start + 2}:$R${start + total + 1}"]
    range_list.append(condi_FSC)
    
    condi_NPS = [f"S{start + 2}:U{start + total + 1}",f"S{start + 2}:S{start + total + 1}", \
                    f"T{start + 2}:T{start + total + 1}",f"U{start + 2}:U{start + total + 1}"]
    range_list.append(condi_NPS)

    condi_vol = [f"V{start + 2}:V{start + total + 1}",f"W{start + 2}:W{start + total + 1}"]
    range_list.append(condi_vol)

    # 9
    add_range_4 = [f"M{start + total*2 + space*2}:R{start + total*2 + space*2}",f"S{start + total*2 + space*2}:U{start + total*2 + space*2}"]
    range_list.append(add_range_4)

    #10 11
    condi_FSC1 = [f"$J{start + total*2 + space*2 + 2}",f"$R${start + total*2 + space*2 + 2}:$R${start + total*2 + space*2 + fsnumber + 1}"]
    range_list.append(condi_FSC1)
    
    condi_NPS1 = [f"S{start + total*2 + space*2 + 2}:U{start + total*2 + space*2 + fsnumber + 1}", \
                  f"S{start + total*2 + space*2 + 2}:S{start + total*2 + space*2 + fsnumber + 1}", \
                    f"T{start + total*2 + space*2 + 2}:T{start + total*2 + space*2 + fsnumber + 1}", \
                        f"U{start + total*2 + space*2 + 2}:U{start + total*2 + space*2 + fsnumber + 1}"]
    range_list.append(condi_NPS1)

    
    #print("Start DataFrame Range:", start_range)
    # print("P_R DataFrame Range:", range_list[0][0])
    # print("I_R DataFrame Range:", range_list[0][1])

    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list,mxndataframe):
    # format the result table

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    print("Information Collection ...")
    sheet = wb["Start"]
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value

    # populate first page information
    start_page_info(sheet,mxndataframe)
    
    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    apply_border(sheet,range_list)
    apply_title(sheet,range_list)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[1]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00'
    
    #for range_name in range_list[7][0]:
    for range_name in [range_list[7][0], range_list[11][0]]:
        for row in sheet[range_name]:
            for cell in row:
                #cell.style = NamedStyle(name='decimal_style', number_format='0.00')
                cell.number_format = '0.00%'


    print("Apply Font ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(italic=True, size = 10)
    
    print("Apply Shading ...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[2]+range_list[3]+range_list[4]+range_list[5]+range_list[9]:
        for row in sheet[range_name]:
            for cell in row:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    condtional_formating(sheet,range_list)

    print("Apply Column Length ...")
    total = int(range_list[0][0][-2:])
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, total+1))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space
    max_length = max(len(str(sheet.cell(row=i, column=3).value)) for i in range(2, total+1))
    sheet.column_dimensions['H'].width = max_length + 2  # Add a little extra space

    # Auto-size columns after applying formatting
    for col_letter in ["C","D","E","F","G","H","I","J","K","L"]:
        sheet.column_dimensions[col_letter].auto_size = True

    # # Save changes
    # wb.save(excel_file)
    return

def start_page_info(sheet,mxndataframe):
    
    # populate voltage information
    index = 11
    column = 16
    while True:
        node = sheet.cell(row=index,column=column).value
        if node is not None:
            # Filter mxndataframe for the current feeder and node    
            node_data = mxndataframe[(mxndataframe['Node'] == node)]
            # Check if node_data is not empty
            if not node_data.empty:
                # get the minimum voltage
                sheet.cell(row=index, column=column + 1).value = node_data['Minimum Voltage (kV)'].iloc[0]
            else:
                # Handle the case when node_data is empty (no matching rows found)
                print(f"No data found for node {node}")
        index = index + 1
        check = sheet.cell(row=index,column=column).value
        check1 = sheet.cell(row=index+1,column=column).value
        if check is None and check1 is None:
            break
    
    return

def apply_border(sheet,range_list):
    print("Apply Border ...")
    for range_name in range_list[0]+range_list[2]+range_list[3]+range_list[4]+range_list[5]+range_list[9]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))

                # Align cell content to the middle
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    return

def apply_title(sheet,range_list):

    print("Apply Title ...")
    for range_name in range_list[3]:
        sheet.merge_cells(range_name)
        part = range_name.split(":")
        cell = sheet[part[0]]
        cell.value = "Site Information Summary"
        cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][0])
    part = range_list[4][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Maximum Average Apparent Power (MVA)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][1])
    part = range_list[4][1].split(":")
    cell = sheet[part[0]]
    cell.value = "Peak NPS (%)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[4][2])
    part = range_list[4][2].split(":")
    cell = sheet[part[0]]
    cell.value = "Voltage (kV)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[5][0])
    part = range_list[5][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Maximum Incoming Feeder RMS Current (A)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[9][0])
    part = range_list[9][0].split(":")
    cell = sheet[part[0]]
    cell.value = "Maximum Average Apparent Power (MVA)"
    cell.font = Font(bold=True)

    sheet.merge_cells(range_list[9][1])
    part = range_list[9][1].split(":")
    cell = sheet[part[0]]
    cell.value = "Peak NPS (%)"
    cell.font = Font(bold=True)

    return

def condtional_formating(sheet,range_list):
    print("Apply Conditional Formatting ...")
    # Compare values in columns I and J for each row and shade accordingly
    # set the pattern fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    # power
    #yellow_line = str(float(sheet[range_list[6][0]])*0.9)
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'lessThan',formula=[range_list[6][0]+'*0.9'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[6][0]],fill=red_fill))
    sheet.conditional_formatting.add(range_list[6][1],CellIsRule(operator = 'between',formula=[range_list[6][0]+'*0.9',range_list[6][0]],fill=yellow_fill))
    # 1 min NPS
    sheet.conditional_formatting.add(range_list[7][1],CellIsRule(operator = 'lessThan',formula=['0.018'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[7][1],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.02'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[7][1],CellIsRule(operator = 'between',formula=['0.018','0.02'],fill=yellow_fill))
    # 10 min NPS
    sheet.conditional_formatting.add(range_list[7][2],CellIsRule(operator = 'lessThan',formula=['0.0135'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[7][2],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.015'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[7][2],CellIsRule(operator = 'between',formula=['0.0135','0.015'],fill=yellow_fill))
    # 30 min NPS
    sheet.conditional_formatting.add(range_list[7][3],CellIsRule(operator = 'lessThan',formula=['0.009'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[7][3],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.01'],fill=red_fill))
    sheet.conditional_formatting.add(range_list[7][3],CellIsRule(operator = 'between',formula=['0.009','0.01'],fill=yellow_fill))
    # Min V
    sheet.conditional_formatting.add(range_list[8][0],CellIsRule(operator = 'greaterThan',formula=['22.5'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[8][0],CellIsRule(operator = 'between',formula=['19','22.5'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[8][0],CellIsRule(operator = 'lessThanOrEqual',formula=['19'],fill=red_fill))
    # Max V
    sheet.conditional_formatting.add(range_list[8][1],CellIsRule(operator = 'lessThan',formula=['27.5'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[8][1],CellIsRule(operator = 'greaterThanOrEqual',formula=['27.5'],fill=red_fill))

    # power
    #yellow_line = str(float(sheet[range_list[10][0]])*0.9)
    sheet.conditional_formatting.add(range_list[10][1],CellIsRule(operator = 'lessThan',formula=[range_list[10][0]+'*0.9'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[10][1],CellIsRule(operator = 'between',formula=[range_list[10][0]+'*0.9',range_list[10][0]],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[10][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[10][0]],fill=red_fill))
    # 1 min NPS
    sheet.conditional_formatting.add(range_list[11][1],CellIsRule(operator = 'lessThan',formula=['0.018'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[11][1],CellIsRule(operator = 'between',formula=['0.018','0.02'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[11][1],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.02'],fill=red_fill))
    # 10 min NPS
    sheet.conditional_formatting.add(range_list[11][2],CellIsRule(operator = 'lessThan',formula=['0.0135'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[11][2],CellIsRule(operator = 'between',formula=['0.0135','0.015'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[11][2],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.015'],fill=red_fill))
    # 30 min NPS
    sheet.conditional_formatting.add(range_list[11][3],CellIsRule(operator = 'lessThan',formula=['0.009'],fill=green_fill))
    sheet.conditional_formatting.add(range_list[11][3],CellIsRule(operator = 'between',formula=['0.009','0.01'],fill=yellow_fill))
    sheet.conditional_formatting.add(range_list[11][3],CellIsRule(operator = 'greaterThanOrEqual',formula=['0.01'],fill=red_fill))

    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "052_N-0_New"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "3"  # Adjust as needed
    text_input = "output"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = None  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

